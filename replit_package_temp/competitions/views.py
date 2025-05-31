from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseNotAllowed, Http404, HttpResponse
from django.utils import timezone
from django.contrib import messages
from django.db.models import Count, Sum, Avg, F, Q, Case, When, IntegerField, Max
from django.views.decorators.http import require_POST
from django.views.decorators.cache import cache_page
from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache
from django.urls import reverse
from django.core.exceptions import ValidationError
from django.db import transaction

from datetime import timedelta
import random
import math
import logging
import json
from io import BytesIO

# Excel imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# PDF imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import MathQuestion, Competition, UserResponse, CompetitionResult, Participant, StudentSession, StudentResponse
from accounts.models import Profile

logger = logging.getLogger('competitions')

def home(request):
    """Home page view"""
    # Get some statistics for the home page
    total_competitions = Competition.objects.count()
    total_users = Profile.objects.filter(user_type='student').count()
    recent_competitions = Competition.objects.filter(is_completed=True).order_by('-end_time')[:5]

    context = {
        'total_competitions': total_competitions,
        'total_users': total_users,
        'recent_competitions': recent_competitions,
    }
    return render(request, 'competitions/home.html', context)

def about(request):
    """About page view"""
    return render(request, 'competitions/about.html')

@login_required
def start_competition(request):
    """Start a new competition"""
    if request.method == 'POST':
        # Get the selected difficulty level
        difficulty = int(request.POST.get('difficulty', 1))
        if difficulty < 1 or difficulty > 9:
            difficulty = 1

        # Get the participant
        participant_id = request.POST.get('participant_name')
        if participant_id:
            participant = get_object_or_404(Participant, id=participant_id)
        else:
            messages.error(request, 'يرجى اختيار مشارك')
            # Get selected grade level or all participants
            grade_filter = request.session.get('grade_filter', None)
            if grade_filter and grade_filter != 'all':
                participants = Participant.objects.filter(grade=grade_filter).order_by('name')
            else:
                participants = Participant.objects.all().order_by('name')

            grades = Participant.GRADE_CHOICES
            return render(request, 'competitions/start.html', {
                'participants': participants,
                'grades': grades,
                'current_grade': grade_filter
            })

        # Create a new competition
        competition = Competition.objects.create(
            user=request.user,
            difficulty=difficulty,
            participant=participant
        )

        # Generate questions for this competition
        generate_competition_questions(competition)

        # Redirect to the first question
        return redirect('competitions:question', question_number=1)

    # Store grade filter in session if provided
    grade_filter = request.GET.get('grade')
    if grade_filter:
        request.session['grade_filter'] = grade_filter
    else:
        grade_filter = request.session.get('grade_filter', None)

    # Get participants based on grade filter (only if grade is selected)
    participants = []
    if grade_filter and grade_filter != 'all':
        participants = Participant.objects.filter(grade=grade_filter).order_by('name')

    # Get all grade choices for the filter dropdown
    grades = Participant.GRADE_CHOICES

    # Get grade statistics for display
    grade_stats = {}
    total_participants = 0
    for grade_value, grade_display in grades:
        count = Participant.objects.filter(grade=grade_value).count()
        if count > 0:
            grade_stats[grade_value] = {
                'display': grade_display,
                'count': count
            }
            total_participants += count

    return render(request, 'competitions/start.html', {
        'participants': participants,
        'grades': grades,
        'grade_stats': grade_stats,
        'total_participants': total_participants,
        'current_grade': grade_filter
    })

@login_required
def question(request, question_number):
    """Display a competition question"""
    # Get the user's active competition
    active_competition = get_active_competition(request.user)
    if not active_competition:
        messages.error(request, 'لا توجد مسابقة نشطة')
        return redirect('competitions:start_competition')

    # Make sure the question number is valid (1-15)
    if question_number < 1 or question_number > 15:
        return redirect('competitions:question', question_number=1)

    # Get the questions for this competition
    responses = active_competition.responses.all().order_by('id')

    # If we have fewer than 15 questions, something went wrong
    if responses.count() < 15:
        messages.error(request, 'حدث خطأ في إعداد المسابقة')
        return redirect('competitions:start_competition')

    # Get the current question
    current_response = responses[question_number-1]
    current_question = current_response.question

    # Check if we should show the next question button or the finish button
    is_last_question = question_number == 15

    # Set time limit for questions based on difficulty level
    time_limit = MathQuestion.get_time_per_question(active_competition.difficulty)

    context = {
        'competition': active_competition,
        'question': current_question,
        'question_number': question_number,
        'response': current_response,
        'is_last_question': is_last_question,
        'time_limit': time_limit,
    }
    return render(request, 'competitions/question.html', context)

@require_POST
@login_required
def submit_answer(request, question_number):
    """Submit an answer to a question"""
    # Get the user's active competition
    active_competition = get_active_competition(request.user)
    if not active_competition:
        return JsonResponse({'status': 'error', 'message': 'لا توجد مسابقة نشطة'})

    # Make sure the question number is valid
    if question_number < 1 or question_number > 15:
        return JsonResponse({'status': 'error', 'message': 'رقم السؤال غير صالح'})

    # Get the user's answer
    user_answer = request.POST.get('answer', '')
    response_time = float(request.POST.get('response_time', 15.0))

    # Get time limit for this difficulty level
    time_limit = MathQuestion.get_time_per_question(active_competition.difficulty)

    # Ensure response time is within limits
    if response_time < 0 or response_time > time_limit:
        response_time = float(time_limit)

    # Get the response object for this question
    responses = active_competition.responses.all().order_by('id')
    if question_number > len(responses):
        return JsonResponse({'status': 'error', 'message': 'رقم السؤال غير صالح'})

    response = responses[question_number-1]
    question = response.question

    # Convert user_answer to float for comparison, if provided
    is_correct = False
    if user_answer and user_answer.strip():
        try:
            user_answer_float = float(user_answer)

            # For division, allow small floating point differences
            if question.operation == 'division':
                is_correct = abs(user_answer_float - question.answer) < 0.001
            else:
                is_correct = user_answer_float == question.answer
        except ValueError:
            is_correct = False

    # Update the response
    response.user_answer = float(user_answer) if user_answer and user_answer.strip() else None
    response.is_correct = is_correct
    response.response_time = response_time
    response.save()

    # Check if this is the last question
    if question_number == 15:
        # Complete the competition
        active_competition.is_completed = True
        active_competition.end_time = timezone.now()
        active_competition.save()

        # Generate the results
        generate_competition_results(active_competition)

        # Return the redirect URL to the results page
        redirect_url = reverse('competitions:results', kwargs={'competition_id': active_competition.id})
        return JsonResponse({
            'status': 'success',
            'is_correct': is_correct,
            'correct_answer': question.answer,
            'is_last_question': True,
            'redirect_url': redirect_url
        })

    # Return the result
    next_question = question_number + 1
    next_url = reverse('competitions:question', kwargs={'question_number': next_question})

    # Check if this is the last question
    is_last_question = question_number == 15

    # Set redirect URL if it's the last question
    redirect_url = reverse('competitions:results', kwargs={'competition_id': active_competition.id}) if is_last_question else None

    return JsonResponse({
        'status': 'success',
        'is_correct': is_correct,
        'correct_answer': question.answer,
        'is_last_question': is_last_question,
        'next_url': next_url,
        'redirect_url': redirect_url
    })

@login_required
def results(request, competition_id):
    """Display the results of a competition"""
    # Get the competition
    competition = get_object_or_404(Competition, id=competition_id, user=request.user)

    # Make sure the competition is completed
    if not competition.is_completed:
        messages.error(request, 'هذه المسابقة لم تكتمل بعد')
        return redirect('competitions:start_competition')

    # Get the competition result
    try:
        result = competition.result
    except CompetitionResult.DoesNotExist:
        # If the result doesn't exist, generate it
        result = generate_competition_results(competition)

    # Get all responses for this competition
    responses = competition.responses.all().order_by('id')

    context = {
        'competition': competition,
        'result': result,
        'responses': responses,
    }
    return render(request, 'competitions/results.html', context)

@login_required
def competition_history(request):
    """Display organized competition history by grade and difficulty"""
    # Get filter parameters
    participant_filter = request.GET.get('participant', '')
    operation_filter = request.GET.get('operation', '')
    min_score = request.GET.get('min_score', '')
    max_score = request.GET.get('max_score', '')
    max_time = request.GET.get('max_time', '')
    grade_filter = request.GET.get('grade', '')
    difficulty_filter = request.GET.get('difficulty', '')

    # Get sorting parameters
    sort_by = request.GET.get('sort_by', 'date_newest')  # Default sorting

    # Base queryset
    competitions = Competition.objects.filter(
        user=request.user,
        is_completed=True
    ).select_related('participant', 'result').prefetch_related('responses__question')

    # Apply filters
    if participant_filter:
        competitions = competitions.filter(participant__name__icontains=participant_filter)

    if operation_filter:
        competitions = competitions.filter(responses__question__operation=operation_filter).distinct()

    if grade_filter:
        competitions = competitions.filter(participant__grade=grade_filter)

    if difficulty_filter:
        competitions = competitions.filter(difficulty=difficulty_filter)

    if min_score:
        try:
            min_score_val = int(min_score)
            competitions = competitions.filter(result__total_score__gte=min_score_val)
        except ValueError:
            pass

    if max_score:
        try:
            max_score_val = int(max_score)
            competitions = competitions.filter(result__total_score__lte=max_score_val)
        except ValueError:
            pass

    if max_time:
        try:
            max_time_val = float(max_time)
            # Calculate average response time for each competition
            competitions = competitions.annotate(
                avg_response_time=Avg('responses__response_time')
            ).filter(avg_response_time__lte=max_time_val)
        except ValueError:
            pass

    # Organize competitions by grade and difficulty
    organized_competitions = {}

    for grade_value, grade_display in Participant.GRADE_CHOICES:
        grade_competitions = competitions.filter(participant__grade=grade_value)

        if grade_competitions.exists():
            organized_competitions[grade_value] = {
                'grade_display': grade_display,
                'difficulties': {}
            }

            # Group by difficulty within each grade
            for difficulty in range(1, 10):  # 9 difficulty levels
                difficulty_competitions = grade_competitions.filter(difficulty=difficulty)

                if difficulty_competitions.exists():
                    # Calculate statistics for each competition
                    competitions_with_stats = []
                    for comp in difficulty_competitions:
                        total_questions = comp.responses.count()
                        correct_answers = comp.responses.filter(is_correct=True).count()
                        avg_time = comp.responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

                        # Calculate duration
                        duration = None
                        if comp.start_time and comp.end_time:
                            duration = comp.end_time - comp.start_time

                        competitions_with_stats.append({
                            'competition': comp,
                            'total_questions': total_questions,
                            'correct_answers': correct_answers,
                            'accuracy': (correct_answers / total_questions * 100) if total_questions > 0 else 0,
                            'avg_response_time': avg_time,
                            'duration': duration,
                            'score': comp.result.total_score if comp.result else 0
                        })

                    # Apply sorting based on sort_by parameter
                    if sort_by == 'score_highest':
                        competitions_with_stats.sort(key=lambda x: x['score'], reverse=True)
                    elif sort_by == 'score_lowest':
                        competitions_with_stats.sort(key=lambda x: x['score'])
                    elif sort_by == 'time_fastest':
                        competitions_with_stats.sort(key=lambda x: x['avg_response_time'])
                    elif sort_by == 'time_slowest':
                        competitions_with_stats.sort(key=lambda x: x['avg_response_time'], reverse=True)
                    elif sort_by == 'date_oldest':
                        competitions_with_stats.sort(key=lambda x: x['competition'].end_time or x['competition'].start_time)
                    else:  # date_newest (default)
                        competitions_with_stats.sort(key=lambda x: x['competition'].end_time or x['competition'].start_time, reverse=True)

                    organized_competitions[grade_value]['difficulties'][difficulty] = competitions_with_stats

    # Get filter options for dropdowns
    all_participants = Participant.objects.filter(
        competitions__user=request.user
    ).distinct().order_by('name')

    operations = [
        ('addition', 'الجمع'),
        ('subtraction', 'الطرح'),
        ('multiplication', 'الضرب'),
        ('division', 'القسمة')
    ]

    # Sorting options
    sort_options = [
        ('date_newest', '📅 التاريخ (الأحدث أولاً)'),
        ('date_oldest', '📅 التاريخ (الأقدم أولاً)'),
        ('score_highest', '🔼 أعلى نقطة'),
        ('score_lowest', '🔽 أقل نقطة'),
        ('time_fastest', '🕒 أسرع وقت للإجابة'),
        ('time_slowest', '🕓 أبطأ وقت للإجابة'),
    ]

    context = {
        'organized_competitions': organized_competitions,
        'all_participants': all_participants,
        'operations': operations,
        'grade_choices': Participant.GRADE_CHOICES,
        'difficulty_choices': [(i, f'مستوى {i}') for i in range(1, 10)],
        'sort_options': sort_options,
        'current_sort': sort_by,
        'filters': {
            'participant': participant_filter,
            'operation': operation_filter,
            'min_score': min_score,
            'max_score': max_score,
            'max_time': max_time,
            'grade': grade_filter,
            'difficulty': difficulty_filter,
            'sort_by': sort_by,
        },
        'has_competitions': competitions.exists(),
    }
    return render(request, 'competitions/history.html', context)

@login_required
def export_history_excel(request):
    """Export competition history as Excel file"""
    import pandas as pd
    from django.http import HttpResponse
    import io

    try:
        # Get competitions in batches to avoid SQLite expression tree limit
        competitions = Competition.objects.filter(
            user=request.user,
            is_completed=True
        ).select_related('participant', 'result').order_by('-end_time')

        # Limit to recent competitions to avoid performance issues
        competitions = competitions[:100]  # Last 100 competitions

        # Prepare data for Excel
        history_data = []

        # Process competitions in smaller batches
        batch_size = 10
        for i in range(0, len(competitions), batch_size):
            batch = competitions[i:i + batch_size]

            for comp in batch:
                try:
                    # Get basic statistics without complex prefetch
                    responses = UserResponse.objects.filter(competition=comp)
                    total_questions = responses.count()
                    correct_answers = responses.filter(is_correct=True).count()

                    # Calculate average time safely
                    avg_time_result = responses.aggregate(avg_time=Avg('response_time'))
                    avg_time = avg_time_result['avg_time'] or 0

                    # Get operation breakdown with simpler queries
                    operations_breakdown = {}
                    for operation in ['addition', 'subtraction', 'multiplication', 'division']:
                        try:
                            op_responses = responses.filter(question__operation=operation)
                            op_correct = op_responses.filter(is_correct=True).count()
                            op_total = op_responses.count()
                            operations_breakdown[operation] = f"{op_correct}/{op_total}" if op_total > 0 else "0/0"
                        except Exception:
                            operations_breakdown[operation] = "0/0"

                    # Calculate duration
                    duration = ""
                    if comp.start_time and comp.end_time:
                        duration_delta = comp.end_time - comp.start_time
                        minutes = duration_delta.total_seconds() // 60
                        seconds = duration_delta.total_seconds() % 60
                        duration = f"{int(minutes):02d}:{int(seconds):02d}"

                    history_data.append({
                        'اسم المشارك': comp.participant.name if comp.participant else 'غير محدد',
                        'المستوى الدراسي': comp.participant.get_grade_display() if comp.participant else 'غير محدد',
                        'مستوى الصعوبة': f"مستوى {comp.difficulty}",
                        'تاريخ البدء': comp.start_time.strftime('%Y-%m-%d %H:%M') if comp.start_time else '',
                        'تاريخ الانتهاء': comp.end_time.strftime('%Y-%m-%d %H:%M') if comp.end_time else '',
                        'النتيجة النهائية': comp.result.total_score if comp.result else 0,
                        'الأسئلة الصحيحة': correct_answers,
                        'إجمالي الأسئلة': total_questions,
                        'نسبة النجاح (%)': round((correct_answers / total_questions * 100), 2) if total_questions > 0 else 0,
                        'متوسط وقت الإجابة (ثانية)': round(avg_time, 2),
                        'المدة الإجمالية': duration,
                        'الجمع': operations_breakdown['addition'],
                        'الطرح': operations_breakdown['subtraction'],
                        'الضرب': operations_breakdown['multiplication'],
                        'القسمة': operations_breakdown['division']
                    })
                except Exception as e:
                    # Skip problematic competitions
                    print(f"خطأ في معالجة المسابقة {comp.id}: {str(e)}")
                    continue

        # Create Excel file
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main history sheet
            if history_data:
                history_df = pd.DataFrame(history_data)
                history_df.to_excel(writer, sheet_name='سجل المسابقات', index=False)
            else:
                # Create empty sheet with message
                empty_df = pd.DataFrame([{'رسالة': 'لا توجد مسابقات للتصدير'}])
                empty_df.to_excel(writer, sheet_name='سجل المسابقات', index=False)

            # Summary by grade sheet
            grade_summary = []
            for grade_value, grade_display in Participant.GRADE_CHOICES:
                try:
                    grade_competitions = competitions.filter(participant__grade=grade_value)
                    if grade_competitions.exists():
                        total_comps = grade_competitions.count()
                        avg_score_result = grade_competitions.aggregate(avg_score=Avg('result__total_score'))
                        avg_score = avg_score_result['avg_score'] or 0

                        grade_summary.append({
                            'المستوى الدراسي': grade_display,
                            'عدد المسابقات': total_comps,
                            'متوسط النتيجة': round(avg_score, 2)
                        })
                except Exception:
                    # Skip problematic grades
                    continue

            if grade_summary:
                grade_summary_df = pd.DataFrame(grade_summary)
                grade_summary_df.to_excel(writer, sheet_name='ملخص المستويات', index=False)

        # Prepare response
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="سجل_المسابقات.xlsx"'

        return response

    except Exception as e:
        # Return error response
        error_output = io.BytesIO()
        error_data = [{
            'خطأ': f'فشل في تصدير سجل المسابقات: {str(e)}',
            'تاريخ الخطأ': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'نصيحة': 'حاول تقليل عدد المسابقات أو اتصل بالدعم الفني'
        }]
        error_df = pd.DataFrame(error_data)
        error_df.to_excel(error_output, index=False, sheet_name='خطأ')
        error_output.seek(0)

        response = HttpResponse(error_output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_تصدير_السجل.xlsx"'
        return response

@login_required
def export_history_pdf(request):
    """Export competition history as PDF file with Arabic support"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    import io
    import os

    try:
        # Register Arabic font (try to use system fonts or fallback)
        try:
            # Try to register a system Arabic font
            font_paths = [
                'C:/Windows/Fonts/arial.ttf',  # Windows
                '/System/Library/Fonts/Arial.ttf',  # macOS
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            ]

            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Arabic', font_path))
                    font_registered = True
                    break

            if not font_registered:
                # Fallback to default font
                arabic_font = 'Helvetica'
        except Exception:
            arabic_font = 'Helvetica'
        else:
            arabic_font = 'Arabic'

        # Get competitions
        competitions = Competition.objects.filter(
            user=request.user,
            is_completed=True
        ).select_related('participant', 'result').order_by('-end_time')[:20]  # Last 20 competitions

        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="competition_history.pdf"'

        # Create a file-like buffer to receive PDF data
        buffer = io.BytesIO()

        # Create the PDF object using the buffer as its "file"
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Define custom styles for Arabic
        styles = getSampleStyleSheet()

        # Custom title style
        title_style = ParagraphStyle(
            'ArabicTitle',
            parent=styles['Title'],
            fontName=arabic_font,
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=30
        )

        # Custom normal style
        normal_style = ParagraphStyle(
            'ArabicNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_RIGHT
        )

        # Add title (using English as fallback)
        title = Paragraph("Competition History Report / سجل المسابقات", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Prepare table data (using English headers to avoid font issues)
        table_data = [['Participant', 'Grade', 'Difficulty', 'Score', 'Date']]

        for comp in competitions:
            table_data.append([
                comp.participant.name if comp.participant else 'Unknown',
                comp.participant.get_grade_display() if comp.participant else 'Unknown',
                f"Level {comp.difficulty}",  # Use English to avoid font issues
                f"{comp.result.total_score}/45" if comp.result else "0/45",
                comp.end_time.strftime('%Y-%m-%d') if comp.end_time else ''
            ])

        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    except Exception as e:
        # Return error response
        error_response = HttpResponse(content_type='application/pdf')
        error_response['Content-Disposition'] = 'attachment; filename="error_report.pdf"'

        # Create simple error PDF
        error_buffer = io.BytesIO()
        error_doc = SimpleDocTemplate(error_buffer, pagesize=A4)
        error_elements = []

        styles = getSampleStyleSheet()
        error_title = Paragraph(f"Error generating PDF: {str(e)}", styles['Title'])
        error_elements.append(error_title)

        error_doc.build(error_elements)
        error_pdf = error_buffer.getvalue()
        error_buffer.close()
        error_response.write(error_pdf)

        return error_response

@login_required
def repeat_competition(request, competition_id):
    """Repeat a competition with the same settings"""
    # Get the original competition
    original_competition = get_object_or_404(Competition, id=competition_id, user=request.user)

    # Create a new competition with the same settings
    new_competition = Competition.objects.create(
        user=request.user,
        difficulty=original_competition.difficulty,
        participant=original_competition.participant
    )

    # Generate questions for this competition
    generate_competition_questions(new_competition)

    messages.success(request, f'تم إنشاء مسابقة جديدة بنفس إعدادات المسابقة السابقة')

    # Redirect to the first question
    return redirect('competitions:question', question_number=1)

# تم حذف دالة student_stats لتجنب التكرار مع التحليلات المتقدمة
# يمكن للمستخدمين الوصول للإحصائيات من خلال صفحة التحليلات المتقدمة

@login_required
def advanced_analytics(request):
    """Advanced analytics dashboard with comprehensive statistics"""
    # Grade-level statistics
    grade_stats = []
    for grade_value, grade_display in Participant.GRADE_CHOICES:
        participants = Participant.objects.filter(grade=grade_value)
        competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
        responses = UserResponse.objects.filter(competition__in=competitions)

        total_participants = participants.count()
        total_competitions = competitions.count()
        total_correct = responses.filter(is_correct=True).count()
        total_questions = responses.count()

        # Calculate average response time
        avg_response_time = responses.aggregate(
            avg_time=Avg('response_time')
        )['avg_time'] or 0

        # Calculate success rate
        success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0

        grade_stats.append({
            'grade': grade_value,
            'grade_display': grade_display,
            'participant_count': total_participants,
            'total_competitions': total_competitions,
            'avg_response_time': round(avg_response_time, 2),
            'success_rate': round(success_rate, 2),
            'total_questions': total_questions,
            'total_correct': total_correct
        })

    # Operations analysis
    operations_analysis = []
    operation_names = {
        'addition': 'الجمع',
        'subtraction': 'الطرح',
        'multiplication': 'الضرب',
        'division': 'القسمة'
    }

    for operation, operation_name in operation_names.items():
        responses = UserResponse.objects.filter(question__operation=operation)
        total_questions = responses.count()
        correct_answers = responses.filter(is_correct=True).count()

        success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0
        avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

        # Analyze strengths and weaknesses
        strengths, weaknesses = analyze_operation_patterns(operation, responses)

        operations_analysis.append({
            'operation': operation,
            'operation_name': operation_name,
            'questions_count': total_questions,
            'success_rate': round(success_rate, 2),
            'avg_time': round(avg_time, 2),
            'strengths': strengths,
            'weaknesses': weaknesses
        })

    # Time-based performance analytics
    time_analytics = get_time_based_analytics()

    # Top performers
    top_performers = get_top_performers()

    context = {
        'grade_stats': grade_stats,
        'operations_analysis': operations_analysis,
        'time_analytics': time_analytics,
        'top_performers': top_performers,
    }
    return render(request, 'competitions/advanced_analytics.html', context)

@login_required
def participant_profile(request, participant_id):
    """Individual participant analytics profile"""
    participant = get_object_or_404(Participant, id=participant_id)

    # Get all competitions for this participant
    competitions = Competition.objects.filter(
        participant=participant,
        is_completed=True
    ).order_by('-end_time')

    # Calculate individual statistics
    total_competitions = competitions.count()
    responses = UserResponse.objects.filter(competition__in=competitions)
    total_questions = responses.count()
    correct_answers = responses.filter(is_correct=True).count()

    overall_accuracy = (correct_answers / total_questions * 100) if total_questions > 0 else 0
    avg_response_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

    # Operation-specific performance
    operation_performance = []
    operation_names = {
        'addition': 'الجمع',
        'subtraction': 'الطرح',
        'multiplication': 'الضرب',
        'division': 'القسمة'
    }

    for operation, operation_name in operation_names.items():
        op_responses = responses.filter(question__operation=operation)
        if op_responses.exists():
            op_correct = op_responses.filter(is_correct=True).count()
            op_total = op_responses.count()
            op_accuracy = (op_correct / op_total * 100)
            op_avg_time = op_responses.aggregate(avg_time=Avg('response_time'))['avg_time']

            # Analyze strengths and weaknesses for this participant
            strengths, weaknesses = analyze_operation_patterns(operation, op_responses)

            operation_performance.append({
                'operation': operation,
                'operation_name': operation_name,
                'accuracy': round(op_accuracy, 2),
                'avg_time': round(op_avg_time, 2),
                'total_questions': op_total,
                'correct_answers': op_correct,
                'strengths': strengths,
                'weaknesses': weaknesses
            })

    # Performance over time (last 10 competitions)
    recent_competitions = competitions[:10]
    performance_trend = []
    for comp in recent_competitions:
        comp_responses = comp.responses.all()
        comp_accuracy = (comp_responses.filter(is_correct=True).count() / comp_responses.count() * 100) if comp_responses.exists() else 0
        comp_avg_time = comp_responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

        performance_trend.append({
            'date': comp.end_time.strftime('%Y-%m-%d'),
            'accuracy': round(comp_accuracy, 2),
            'avg_time': round(comp_avg_time, 2),
            'difficulty': comp.get_difficulty_display()
        })

    # Recommendations
    recommendations = generate_recommendations(participant, operation_performance)

    context = {
        'participant': participant,
        'total_competitions': total_competitions,
        'overall_accuracy': round(overall_accuracy, 2),
        'avg_response_time': round(avg_response_time, 2),
        'operation_performance': operation_performance,
        'performance_trend': performance_trend,
        'recommendations': recommendations,
        'recent_competitions': recent_competitions[:5]
    }
    return render(request, 'competitions/participant_profile.html', context)

def generate_recommendations(participant, operation_performance):
    """Generate personalized recommendations for a participant"""
    recommendations = []

    for op in operation_performance:
        if op['accuracy'] < 70:
            if op['operation'] == 'multiplication':
                recommendations.append({
                    'type': 'improvement',
                    'title': f"تحسين {op['operation_name']}",
                    'description': f"ينصح بمراجعة جداول الضرب والتدرب عليها يوم",
                    'priority': 'high'
                })
            elif op['operation'] == 'division':
                recommendations.append({
                    'type': 'improvement',
                    'title': f"تحسين {op['operation_name']}",
                    'description': f"ينصح بالتدرب على القسمة البسيطة أولاً ثم التدرج للأصعب",
                    'priority': 'high'
                })
            else:
                recommendations.append({
                    'type': 'improvement',
                    'title': f"تحسين {op['operation_name']}",
                    'description': f"ينصح بالمزيد من التدريب على عمليات {op['operation_name']}",
                    'priority': 'medium'
                })

        elif op['accuracy'] >= 90:
            recommendations.append({
                'type': 'strength',
                'title': f"تميز في {op['operation_name']}",
                'description': f"أداء ممتاز! يمكن الانتقال لمستوى أصعب",
                'priority': 'low'
            })

    # Time-based recommendations
    overall_avg_time = sum(op['avg_time'] for op in operation_performance) / len(operation_performance) if operation_performance else 0
    if overall_avg_time > 15:
        recommendations.append({
            'type': 'speed',
            'title': "تحسين السرعة",
            'description': "ينصح بالتدرب على الحساب الذهني لتحسين سرعة الإجابة",
            'priority': 'medium'
        })

    return recommendations[:5]  # Limit to top 5 recommendations

@login_required
def export_analytics_pdf(request):
    """Export analytics data as PDF report with improved Arabic support"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io

    try:
        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="analytics_report.pdf"'

        # Create a file-like buffer to receive PDF data
        buffer = io.BytesIO()

        # Create the PDF object using the buffer as its "file"
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Add title (using English to avoid font issues)
        title = Paragraph("Advanced Analytics Report - Math Competition Platform", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Get analytics data
        grade_stats = []
        for grade_value, grade_display in Participant.GRADE_CHOICES:
            participants = Participant.objects.filter(grade=grade_value)
            competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
            responses = UserResponse.objects.filter(competition__in=competitions)

            total_participants = participants.count()
            total_competitions = competitions.count()
            total_correct = responses.filter(is_correct=True).count()
            total_questions = responses.count()

            avg_response_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
            success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0

            grade_stats.append([
                f"Grade {grade_value}",  # Use English
                str(total_participants),
                str(total_competitions),
                f"{avg_response_time:.2f}s",
                f"{success_rate:.2f}%"
            ])

        # Create grade statistics table
        grade_table_data = [['Grade Level', 'Participants', 'Competitions', 'Avg Time', 'Success Rate']]
        grade_table_data.extend(grade_stats)

        grade_table = Table(grade_table_data)
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(Paragraph("Grade Level Statistics", styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(grade_table)
        elements.append(Spacer(1, 20))

        # Build PDF
        doc.build(elements)

        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    except Exception as e:
        # Return error response
        error_response = HttpResponse(content_type='application/pdf')
        error_response['Content-Disposition'] = 'attachment; filename="analytics_error.pdf"'

        # Create simple error PDF
        error_buffer = io.BytesIO()
        error_doc = SimpleDocTemplate(error_buffer, pagesize=A4)
        error_elements = []

        styles = getSampleStyleSheet()
        error_title = Paragraph(f"Error generating analytics PDF: {str(e)}", styles['Title'])
        error_elements.append(error_title)

        error_doc.build(error_elements)
        error_pdf = error_buffer.getvalue()
        error_buffer.close()
        error_response.write(error_pdf)

        return error_response

@login_required
def export_analytics_excel(request):
    """Export analytics data as Excel file"""
    import pandas as pd
    from django.http import HttpResponse
    import io

    # Create Excel writer object
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Grade Statistics
        grade_data = []
        for grade_value, grade_display in Participant.GRADE_CHOICES:
            participants = Participant.objects.filter(grade=grade_value)
            competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
            responses = UserResponse.objects.filter(competition__in=competitions)

            total_participants = participants.count()
            total_competitions = competitions.count()
            total_correct = responses.filter(is_correct=True).count()
            total_questions = responses.count()

            avg_response_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
            success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0

            grade_data.append({
                'المستوى الدراسي': grade_display,
                'عدد المشاركين': total_participants,
                'إجمالي المسابقات': total_competitions,
                'متوسط وقت الإجابة (ثانية)': round(avg_response_time, 2),
                'معدل النجاح (%)': round(success_rate, 2),
                'إجمالي الأسئلة': total_questions,
                'الإجابات الصحيحة': total_correct
            })

        grade_df = pd.DataFrame(grade_data)
        grade_df.to_excel(writer, sheet_name='إحصائيات المستويات', index=False)

        # Operations Analysis
        operations_data = []
        operation_names = {
            'addition': 'الجمع',
            'subtraction': 'الطرح',
            'multiplication': 'الضرب',
            'division': 'القسمة'
        }

        for operation, operation_name in operation_names.items():
            responses = UserResponse.objects.filter(question__operation=operation)
            total_questions = responses.count()
            correct_answers = responses.filter(is_correct=True).count()

            success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0
            avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

            # Get strengths and weaknesses
            strengths, weaknesses = analyze_operation_patterns(operation, responses)

            operations_data.append({
                'العملية': operation_name,
                'عدد الأسئلة': total_questions,
                'معدل النجاح (%)': round(success_rate, 2),
                'متوسط الوقت (ثانية)': round(avg_time, 2),
                'الإجابات الصحيحة': correct_answers,
                'نقاط القوة': ', '.join(strengths),
                'نقاط الضعف': ', '.join(weaknesses)
            })

        operations_df = pd.DataFrame(operations_data)
        operations_df.to_excel(writer, sheet_name='تحليل العمليات', index=False)

        # Top Performers
        top_performers = get_top_performers()

        # Top by accuracy
        accuracy_data = []
        for participant in top_performers['top_by_accuracy']:
            accuracy_data.append({
                'اسم المشارك': participant.name,
                'المستوى الدراسي': participant.get_grade_display(),
                'معدل النجاح (%)': round(participant.accuracy, 2),
                'إجمالي الإجابات': participant.total_responses,
                'الإجابات الصحيحة': participant.correct_responses
            })

        accuracy_df = pd.DataFrame(accuracy_data)
        accuracy_df.to_excel(writer, sheet_name='الأعلى دقة', index=False)

        # Top by speed
        speed_data = []
        for participant in top_performers['top_by_speed']:
            speed_data.append({
                'اسم المشارك': participant.name,
                'المستوى الدراسي': participant.get_grade_display(),
                'متوسط وقت الإجابة (ثانية)': round(participant.avg_response_time, 2),
                'إجمالي الإجابات': participant.total_responses
            })

        speed_df = pd.DataFrame(speed_data)
        speed_df.to_excel(writer, sheet_name='الأسرع', index=False)

        # Participation data
        participation_data = []
        for participant in top_performers['top_by_participation']:
            participation_data.append({
                'اسم المشارك': participant.name,
                'المستوى الدراسي': participant.get_grade_display(),
                'عدد المسابقات': participant.competition_count
            })

        participation_df = pd.DataFrame(participation_data)
        participation_df.to_excel(writer, sheet_name='الأكثر مشاركة', index=False)

    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="تحليلات_متقدمة_شاملة.xlsx"'

    return response

# Helper functions for analytics

def analyze_operation_patterns(operation, responses):
    """Analyze strengths and weaknesses for a specific operation"""
    strengths = []
    weaknesses = []

    if operation == 'multiplication':
        # Analyze multiplication tables performance
        for table in range(2, 13):
            table_responses = responses.filter(
                Q(question__first_number=table) | Q(question__second_number=table)
            )
            if table_responses.exists():
                success_rate = table_responses.filter(is_correct=True).count() / table_responses.count() * 100
                if success_rate >= 80:
                    strengths.append(f"جدول الضرب {table}")
                elif success_rate < 50:
                    weaknesses.append(f"جدول الضرب {table}")

    elif operation == 'addition':
        # Analyze number ranges
        small_numbers = responses.filter(question__first_number__lte=10, question__second_number__lte=10)
        large_numbers = responses.filter(Q(question__first_number__gt=50) | Q(question__second_number__gt=50))

        if small_numbers.exists():
            success_rate = small_numbers.filter(is_correct=True).count() / small_numbers.count() * 100
            if success_rate >= 80:
                strengths.append("الأرقام الصغيرة (1-10)")
            elif success_rate < 50:
                weaknesses.append("الأرقام الصغيرة (1-10)")

        if large_numbers.exists():
            success_rate = large_numbers.filter(is_correct=True).count() / large_numbers.count() * 100
            if success_rate >= 80:
                strengths.append("الأرقام الكبيرة (50+)")
            elif success_rate < 50:
                weaknesses.append("الأرقام الكبيرة (50+)")

    elif operation == 'subtraction':
        # Analyze borrowing operations
        borrowing_responses = responses.filter(
            question__first_number__lt=F('question__second_number')
        )
        if borrowing_responses.exists():
            success_rate = borrowing_responses.filter(is_correct=True).count() / borrowing_responses.count() * 100
            if success_rate >= 80:
                strengths.append("عمليات الاستلاف")
            elif success_rate < 50:
                weaknesses.append("عمليات الاستلاف")

    elif operation == 'division':
        # Analyze division by specific numbers
        easy_division = responses.filter(question__second_number__in=[2, 5, 10])
        hard_division = responses.filter(question__second_number__in=[3, 7, 9])

        if easy_division.exists():
            success_rate = easy_division.filter(is_correct=True).count() / easy_division.count() * 100
            if success_rate >= 80:
                strengths.append("القسمة على الأرقام السهلة (2, 5, 10)")
            elif success_rate < 50:
                weaknesses.append("القسمة على الأرقام السهلة (2, 5, 10)")

        if hard_division.exists():
            success_rate = hard_division.filter(is_correct=True).count() / hard_division.count() * 100
            if success_rate >= 80:
                strengths.append("القسمة على الأرقام الصعبة (3, 7, 9)")
            elif success_rate < 50:
                weaknesses.append("القسمة على الأرقام الصعبة (3, 7, 9)")

    return strengths[:3], weaknesses[:3]  # Limit to top 3 each

def get_time_based_analytics():
    """Get time-based performance analytics"""
    from datetime import datetime, timedelta

    # Get data for the last 30 days
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_responses = UserResponse.objects.filter(
        competition__end_time__gte=thirty_days_ago
    )

    # Average response time by operation
    operation_times = recent_responses.values('question__operation').annotate(
        avg_time=Avg('response_time'),
        count=Count('id')
    ).order_by('question__operation')

    # Fastest and slowest participants
    participant_times = recent_responses.values(
        'competition__participant__name'
    ).annotate(
        avg_time=Avg('response_time'),
        total_responses=Count('id')
    ).filter(total_responses__gte=10).order_by('avg_time')

    fastest_participants = participant_times[:5]
    slowest_participants = participant_times.reverse()[:5]

    # Performance improvement tracking
    improvement_data = []
    for participant in Participant.objects.filter(competitions__isnull=False).distinct()[:10]:
        competitions = participant.competitions.filter(is_completed=True).order_by('start_time')
        if competitions.count() >= 3:
            first_avg = competitions[:3].aggregate(
                avg_time=Avg('responses__response_time')
            )['avg_time']
            last_avg = competitions.reverse()[:3].aggregate(
                avg_time=Avg('responses__response_time')
            )['avg_time']

            if first_avg and last_avg:
                improvement = ((first_avg - last_avg) / first_avg) * 100
                improvement_data.append({
                    'participant': participant.name,
                    'improvement': round(improvement, 2)
                })

    return {
        'operation_times': operation_times,
        'fastest_participants': fastest_participants,
        'slowest_participants': slowest_participants,
        'improvement_data': improvement_data[:10]
    }

def get_top_performers():
    """Get top performing participants"""
    from django.db.models import FloatField

    # Top by accuracy
    top_by_accuracy = Participant.objects.annotate(
        total_responses=Count('competitions__responses'),
        correct_responses=Count('competitions__responses', filter=Q(competitions__responses__is_correct=True)),
        accuracy=Case(
            When(total_responses=0, then=0.0),
            default=F('correct_responses') * 100.0 / F('total_responses'),
            output_field=FloatField()
        )
    ).filter(total_responses__gte=5).order_by('-accuracy')[:10]

    # Top by speed (fastest average response time)
    top_by_speed = Participant.objects.annotate(
        avg_response_time=Avg('competitions__responses__response_time'),
        total_responses=Count('competitions__responses')
    ).filter(total_responses__gte=5).order_by('avg_response_time')[:10]

    # Top by total competitions
    top_by_participation = Participant.objects.annotate(
        competition_count=Count('competitions', filter=Q(competitions__is_completed=True))
    ).filter(competition_count__gt=0).order_by('-competition_count')[:10]

    return {
        'top_by_accuracy': top_by_accuracy,
        'top_by_speed': top_by_speed,
        'top_by_participation': top_by_participation
    }

def get_active_competition(user):
    """Get the user's active (incomplete) competition"""
    try:
        return Competition.objects.filter(
            user=user,
            is_completed=False
        ).latest('start_time')
    except Competition.DoesNotExist:
        return None

def generate_competition_questions(competition):
    """Generate questions for a competition based on difficulty level"""
    from .question_generators import QuestionGenerator
    from .models import MathQuestion

    # Get the distribution for this difficulty level
    distribution = MathQuestion.get_question_distribution(competition.difficulty)

    # Generate questions for each operation type
    for operation, count in distribution.items():
        for _ in range(count):
            # Generate a question based on operation type and difficulty
            question = generate_math_question_new(operation, competition.difficulty)

            # Save the question to the database
            db_question = MathQuestion.objects.create(
                operation=operation,
                difficulty=competition.difficulty,
                first_number=question.get('first_number'),
                second_number=question.get('second_number'),
                question_text=question.get('expression', question.get('question_text')),
                answer=question['answer']
            )

            # Create a user response object for this question
            UserResponse.objects.create(
                competition=competition,
                question=db_question
            )

def generate_math_question(operation, difficulty):
    """Generate a single math question"""
    # Define ranges for each operation and difficulty level according to new requirements
    ranges = {
        'addition': {
            1: {'min': 1, 'max': 20},  # Level 1: Numbers 1-20
            2: {'min': 5, 'max': 30, 'avoid_min': 1, 'avoid_max': 5},  # Level 2: Numbers 5-30, avoid 1-5
            3: {'min': 7, 'max': 40, 'operators': [1, 2, 3, 4, 5]},  # Level 3: Operators 1-5
            4: {'min': 9, 'max': 50, 'avoid_min': 1, 'avoid_max': 8},  # Level 4: Numbers 9-50, avoid 1-8
            6: {'min': 2, 'max': 20},  # Special level 6: Addition and subtraction with numbers 2-20
        },
        'subtraction': {
            1: {'min': 1, 'max': 20},  # Level 1: Numbers 1-20
            2: {'min': 5, 'max': 30, 'avoid_min': 1, 'avoid_max': 5},  # Level 2: Numbers 5-30, avoid 1-5
            3: {'min': 7, 'max': 40, 'operators': [1, 2, 3, 4, 5]},  # Level 3: Operators 1-5
            4: {'min': 9, 'max': 50, 'avoid_min': 1, 'avoid_max': 8},  # Level 4: Numbers 9-50, avoid 1-8
            6: {'min': 2, 'max': 20},  # Special level 6: Addition and subtraction with numbers 2-20
        },
        'multiplication': {
            1: {'tables': [2, 3, 4, 5, 6, 7]},  # Level 1: Multiplication tables 2,3,4,5,6,7
            2: {'min': 3, 'max': 12},  # Level 2: Tables 3-12
            3: {'min': 3, 'max': 12},  # Level 3: Tables 3x3 to 12x12
            4: {'min': 1, 'max': 15},  # Level 4: Tables up to 15x15
            5: {'min': 3, 'max': 12},  # Special level 5: Only multiplication tables 3x3 to 12x12
        },
        'division': {
            1: {'quotients': [2, 3, 4], 'divisor_max': 10},  # Level 1: Quotients of 2,3,4
            2: {'quotients': [2, 3, 4, 5], 'divisor_max': 12},  # Level 2: Quotients of 2,3,4,5
            3: {'quotients': [2, 3, 4, 5], 'divisor_max': 12},  # Level 3: Quotients of 2,3,4,5
            4: {'quotients': [3, 4, 5, 6], 'divisor_max': 15},  # Level 4: Quotients of 3,4,5,6
            7: {'quotients': [2, 3, 4, 5], 'dividend_min': 5, 'dividend_max': 60, 'divisor_max': 30},  # Special level 7: Only division with quotients 2-5 and dividend 5-60
        }
    }

    # Get the range for this operation and difficulty
    range_info = ranges[operation][difficulty]

    # Track last generated questions to avoid repeating operations
    if not hasattr(generate_math_question, 'last_questions'):
        generate_math_question.last_questions = []

    # Function to check if numbers are similar (too close to each other)
    def are_numbers_similar(num1, num2):
        # Consider numbers similar if they are the same
        return num1 == num2

    # Function to check if this question is too similar to recent ones
    def is_question_repeated(operation, first_number, second_number):
        for last_op, last_first, last_second in generate_math_question.last_questions[-5:]:
            if last_op == operation and last_first == first_number and last_second == second_number:
                return True
            # Also consider reverse operations as similar
            if last_op == operation and last_first == second_number and last_second == first_number:
                return True
        return False

    # Generate the question based on the operation
    max_attempts = 20  # Prevent infinite loops
    attempts = 0

    while attempts < max_attempts:
        attempts += 1

        if operation == 'addition':
            if difficulty == 3:  # Level 3: One number should be from 1-5
                operator = random.choice(range_info['operators'])
                num = random.randint(range_info['min'], range_info['max'])
                num1, num2 = operator, num
            elif 'avoid_min' in range_info:  # Levels with numbers to avoid
                num1 = random.randint(range_info['min'], range_info['max'])
                while range_info['avoid_min'] <= num1 <= range_info['avoid_max']:
                    num1 = random.randint(range_info['min'], range_info['max'])

                num2 = random.randint(range_info['min'], range_info['max'])
                while range_info['avoid_min'] <= num2 <= range_info['avoid_max'] or are_numbers_similar(num1, num2):
                    num2 = random.randint(range_info['min'], range_info['max'])
            else:  # Standard range
                num1 = random.randint(range_info['min'], range_info['max'])
                num2 = random.randint(range_info['min'], range_info['max'])
                while are_numbers_similar(num1, num2):
                    num2 = random.randint(range_info['min'], range_info['max'])

            # العدد الأصغر على اليسار والأكبر على اليمين
            first_number = min(num1, num2)
            second_number = max(num1, num2)

            answer = first_number + second_number

        elif operation == 'subtraction':
            if difficulty == 3:  # Level 3: One number should be from 1-5
                operator = random.choice(range_info['operators'])
                num = random.randint(range_info['min'], range_info['max'])
                # Generate two numbers
                num1 = operator
                num2 = num
            elif 'avoid_min' in range_info:  # Levels with numbers to avoid
                num1 = random.randint(range_info['min'], range_info['max'])
                while range_info['avoid_min'] <= num1 <= range_info['avoid_max']:
                    num1 = random.randint(range_info['min'], range_info['max'])

                num2 = random.randint(range_info['min'], range_info['max'])
                while (range_info['avoid_min'] <= num2 <= range_info['avoid_max'] or
                       are_numbers_similar(num1, num2)):
                    num2 = random.randint(range_info['min'], range_info['max'])
            else:
                num1 = random.randint(range_info['min'], range_info['max'])
                num2 = random.randint(range_info['min'], range_info['max'])
                while are_numbers_similar(num1, num2):
                    num2 = random.randint(range_info['min'], range_info['max'])

            # العدد الأصغر على اليسار والأكبر على اليمين
            # لكن للطرح نحتاج النتيجة موجبة، لذا نطرح الأصغر من الأكبر
            smaller_num = min(num1, num2)
            larger_num = max(num1, num2)
            first_number = smaller_num
            second_number = larger_num

            answer = second_number - first_number

        elif operation == 'multiplication':
            if 'tables' in range_info:  # Level 1: Specific multiplication tables
                num1 = random.choice(range_info['tables'])
                num2 = random.randint(1, 10)  # Any number from 1-10 to multiply with
                while are_numbers_similar(num1, num2):
                    num2 = random.randint(1, 10)
            else:
                num1 = random.randint(range_info['min'], range_info['max'])
                num2 = random.randint(range_info['min'], range_info['max'])
                while are_numbers_similar(num1, num2):
                    num2 = random.randint(range_info['min'], range_info['max'])

            # العدد الأصغر على اليسار والأكبر على اليمين
            first_number = min(num1, num2)
            second_number = max(num1, num2)

            answer = first_number * second_number

        elif operation == 'division':
            # For division, we use specific quotients based on the level
            quotient = random.choice(range_info['quotients'])

            if difficulty == 7:  # Special level for division with dividend 5-60
                # For level 7, we need to handle the specific dividend range
                # First try to find divisors that would give a dividend within our range
                max_divisor = min(range_info['divisor_max'], range_info['dividend_max'] // quotient)
                min_divisor = max(2, range_info['dividend_min'] // quotient)

                if min_divisor <= max_divisor:
                    divisor = random.randint(min_divisor, max_divisor)
                    dividend = quotient * divisor

                    # Double-check dividend is in the required range
                    if dividend < range_info['dividend_min'] or dividend > range_info['dividend_max']:
                        # Adjust if somehow outside range
                        divisor = min(max_divisor, range_info['dividend_max'] // quotient)
                        dividend = quotient * divisor
                else:
                    # Fallback if constraints are impossible to satisfy
                    divisor = random.randint(2, range_info['divisor_max'])
                    dividend = quotient * divisor
                    # Cap at maximum allowed dividend
                    if dividend > range_info['dividend_max']:
                        dividend = range_info['dividend_max']
                        divisor = dividend // quotient
                        # Recalculate quotient if needed
                        quotient = dividend // divisor
            else:
                # Regular division for other levels
                divisor = random.randint(2, range_info['divisor_max'])
                dividend = quotient * divisor

            # العدد الأصغر (المقسوم عليه) على اليسار والأكبر (المقسوم) على اليمين
            first_number = divisor
            second_number = dividend
            answer = quotient

        # Check if this question is a repeat of recent ones
        if not is_question_repeated(operation, first_number, second_number):
            break

    # Store this question in the history
    generate_math_question.last_questions.append((operation, first_number, second_number))
    # Keep only the last 10 questions in history
    if len(generate_math_question.last_questions) > 10:
        generate_math_question.last_questions = generate_math_question.last_questions[-10:]

    return {
        'first_number': first_number,
        'second_number': second_number,
        'answer': answer
    }

def generate_math_question_new(operation, difficulty):
    """Generate a math question using the new system"""
    from .question_generators import QuestionGenerator

    if operation == 'addition':
        return QuestionGenerator.generate_addition_question(difficulty)
    elif operation == 'subtraction':
        return QuestionGenerator.generate_subtraction_question(difficulty)
    elif operation == 'multiplication':
        return QuestionGenerator.generate_multiplication_question(difficulty)
    elif operation == 'division':
        return QuestionGenerator.generate_division_question(difficulty)
    elif operation == 'mixed_operations':
        return QuestionGenerator.generate_mixed_operations_question(difficulty)
    elif operation == 'fractions':
        return QuestionGenerator.generate_fractions_question()
    elif operation == 'algebra':
        return QuestionGenerator.generate_algebra_question()
    elif operation == 'geometry':
        return QuestionGenerator.generate_geometry_question()
    elif operation == 'trigonometry':
        return QuestionGenerator.generate_trigonometry_question()
    elif operation == 'word_problems':
        return QuestionGenerator.generate_word_problem(difficulty)
    else:
        # Fallback to basic addition
        return QuestionGenerator.generate_addition_question(1)

def generate_competition_results(competition):
    """Generate results for a completed competition"""
    # Get all responses for this competition
    responses = competition.responses.all()

    # Calculate totals for each operation
    addition_total = responses.filter(question__operation='addition').count()
    subtraction_total = responses.filter(question__operation='subtraction').count()
    multiplication_total = responses.filter(question__operation='multiplication').count()
    division_total = responses.filter(question__operation='division').count()
    mixed_operations_total = responses.filter(question__operation='mixed_operations').count()
    fractions_total = responses.filter(question__operation='fractions').count()
    algebra_total = responses.filter(question__operation='algebra').count()
    geometry_total = responses.filter(question__operation='geometry').count()
    trigonometry_total = responses.filter(question__operation='trigonometry').count()
    word_problems_total = responses.filter(question__operation='word_problems').count()

    # Calculate correct answers for each operation
    addition_correct = responses.filter(question__operation='addition', is_correct=True).count()
    subtraction_correct = responses.filter(question__operation='subtraction', is_correct=True).count()
    multiplication_correct = responses.filter(question__operation='multiplication', is_correct=True).count()
    division_correct = responses.filter(question__operation='division', is_correct=True).count()
    mixed_operations_correct = responses.filter(question__operation='mixed_operations', is_correct=True).count()
    fractions_correct = responses.filter(question__operation='fractions', is_correct=True).count()
    algebra_correct = responses.filter(question__operation='algebra', is_correct=True).count()
    geometry_correct = responses.filter(question__operation='geometry', is_correct=True).count()
    trigonometry_correct = responses.filter(question__operation='trigonometry', is_correct=True).count()
    word_problems_correct = responses.filter(question__operation='word_problems', is_correct=True).count()

    # Calculate total score (3 points per correct answer)
    total_correct = (addition_correct + subtraction_correct + multiplication_correct + division_correct +
                    mixed_operations_correct + fractions_correct + algebra_correct + geometry_correct +
                    trigonometry_correct + word_problems_correct)
    total_score = total_correct * 3

    # Create or update the competition result
    result, created = CompetitionResult.objects.update_or_create(
        competition=competition,
        defaults={
            'total_score': total_score,
            'addition_correct': addition_correct,
            'subtraction_correct': subtraction_correct,
            'multiplication_correct': multiplication_correct,
            'division_correct': division_correct,
            'mixed_operations_correct': mixed_operations_correct,
            'fractions_correct': fractions_correct,
            'algebra_correct': algebra_correct,
            'geometry_correct': geometry_correct,
            'trigonometry_correct': trigonometry_correct,
            'word_problems_correct': word_problems_correct,
            'addition_total': addition_total,
            'subtraction_total': subtraction_total,
            'multiplication_total': multiplication_total,
            'division_total': division_total,
            'mixed_operations_total': mixed_operations_total,
            'fractions_total': fractions_total,
            'algebra_total': algebra_total,
            'geometry_total': geometry_total,
            'trigonometry_total': trigonometry_total,
            'word_problems_total': word_problems_total,
        }
    )

    return result
@login_required
def export_results_excel(request):
    """Export participant results as Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        import logging

        logger = logging.getLogger('competitions')
        logger.info(f"تصدير النتائج بواسطة المستخدم: {request.user.username}")

        # Create a BytesIO buffer to receive the Excel file
        buffer = BytesIO()

        # Get all participants with their competition results
        participants = Participant.objects.annotate(
            competition_count=Count('competitions'),
            average_score=Avg('competitions__result__total_score'),
            best_score=Max('competitions__result__total_score')
        ).filter(competition_count__gt=0).order_by('-average_score')

        if not participants.exists():
            # إنشاء ملف فارغ مع رسالة
            empty_data = [{
                'رسالة': 'لا توجد بيانات للتصدير',
                'تاريخ التصدير': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
            }]
            df = pd.DataFrame(empty_data)
            df.to_excel(buffer, index=False, sheet_name='لا توجد بيانات')
        else:
            # Create Excel data
            data = []
            for participant in participants:
                # Get competition results for this participant
                competitions = Competition.objects.filter(
                    participant=participant,
                    is_completed=True
                ).order_by('-end_time')

                for competition in competitions:
                    try:
                        result = competition.result

                        data.append({
                            'اسم المشارك': participant.name,
                            'الصف الدراسي': participant.get_grade_display(),
                            'رقم المسابقة': competition.id,
                            'تاريخ المسابقة': competition.end_time.strftime('%Y-%m-%d %H:%M'),
                            'المستوى': competition.get_difficulty_display(),
                            'النتيجة الإجمالية': result.total_score,
                            'الجمع (الصحيحة)': result.addition_correct,
                            'الجمع (المجموع)': result.addition_total,
                            'الجمع (النسبة)': result.addition_percentage,
                            'الطرح (الصحيحة)': result.subtraction_correct,
                            'الطرح (المجموع)': result.subtraction_total,
                            'الطرح (النسبة)': result.subtraction_percentage,
                            'الضرب (الصحيحة)': result.multiplication_correct,
                            'الضرب (المجموع)': result.multiplication_total,
                            'الضرب (النسبة)': result.multiplication_percentage,
                            'القسمة (الصحيحة)': result.division_correct,
                            'القسمة (المجموع)': result.division_total,
                            'القسمة (النسبة)': result.division_percentage,
                            'الترتيب': 0  # Will be filled later
                        })
                    except CompetitionResult.DoesNotExist:
                        continue

            # Convert to DataFrame and sort by score
            df = pd.DataFrame(data)
            if not df.empty:
                df = df.sort_values(by='النتيجة الإجمالية', ascending=False)
                df['الترتيب'] = range(1, len(df) + 1)

            # Save to Excel file
            df.to_excel(buffer, index=False, sheet_name='نتائج المسابقات')

        # Set up the response
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # اسم الملف مع التاريخ
        filename = f'نتائج_المسابقات_{timezone.now().strftime("%Y-%m-%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        logger.info(f"تم تصدير النتائج بنجاح: {filename}")
        return response

    except Exception as e:
        logger.error(f"خطأ في تصدير النتائج: {str(e)}")
        # إرجاع ملف خطأ
        error_data = [{
            'خطأ': f'فشل في تصدير البيانات: {str(e)}',
            'تاريخ الخطأ': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        }]
        error_df = pd.DataFrame(error_data)
        error_buffer = BytesIO()
        error_df.to_excel(error_buffer, index=False, sheet_name='خطأ')
        error_buffer.seek(0)

        response = HttpResponse(error_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_التصدير.xlsx"'
        return response

@login_required
def add_participant(request):
    """Add a new participant."""
    if request.method == 'POST':
        # Check if an Excel file was uploaded
        excel_file = request.FILES.get('participant_file')
        if excel_file:
            try:
                import pandas as pd
                import io

                # Read Excel file
                if excel_file.name.endswith('.xlsx'):
                    df = pd.read_excel(excel_file)
                elif excel_file.name.endswith('.csv'):
                    df = pd.read_csv(excel_file)
                else:
                    messages.error(request, 'صيغة الملف غير مدعومة، يرجى استخدام ملف Excel (.xlsx) أو CSV (.csv)')
                    return redirect('competitions:start_competition')

                # Expected columns: name, grade
                if 'name' not in df.columns or 'grade' not in df.columns:
                    messages.error(request, 'يجب أن يحتوي الملف على عمودين: name و grade')
                    return redirect('competitions:start_competition')

                # Add each participant from the Excel file
                participants_added = 0
                for _, row in df.iterrows():
                    try:
                        name = row['name']
                        grade = int(row['grade'])
                        if 1 <= grade <= 9:  # Valid grade range
                            # Use default group 1 for bulk imports
                            group = row.get('group', 1) if 'group' in df.columns else 1
                            try:
                                group = int(group)
                                if group not in [1, 2]:  # Only allow group 1 or 2
                                    group = 1
                            except (ValueError, TypeError):
                                group = 1

                            Participant.objects.create(
                                name=name,
                                grade=grade,
                                group=group
                            )
                            participants_added += 1
                    except (ValueError, TypeError):
                        continue  # Skip invalid rows

                messages.success(request, f'تمت إضافة {participants_added} مشاركين بنجاح')
                return redirect('competitions:start_competition')

            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء معالجة الملف: {str(e)}')
                return redirect('competitions:start_competition')
        else:
            # Check the add_type to determine which form was submitted
            add_type = request.POST.get('add_type', 'legacy')

            if add_type == 'single':
                # New single participant addition
                name = request.POST.get('name', '').strip()
                grade = request.POST.get('grade', '').strip()
                group = request.POST.get('group', '1').strip()

                if not name:
                    messages.error(request, 'الرجاء إدخال اسم المشارك')
                    return redirect('competitions:start_competition')

                if not grade:
                    messages.error(request, 'الرجاء اختيار الصف')
                    return redirect('competitions:start_competition')

                # Convert grade and group to integer and validate
                try:
                    grade = int(grade)
                    if grade < 1 or grade > 9:
                        messages.error(request, 'الصف غير صالح')
                        return redirect('competitions:start_competition')

                    # Validate group
                    group = int(group)
                    if group not in [1, 2]:
                        group = 1

                    # Check if participant already exists
                    if Participant.objects.filter(name=name, grade=grade, group=group).exists():
                        messages.warning(request, f'المشارك {name} موجود بالفعل في نفس الصف والفوج')
                        return redirect('competitions:start_competition')

                    # Create the participant
                    Participant.objects.create(
                        name=name,
                        grade=grade,
                        group=group
                    )

                    messages.success(request, f'✅ تمت إضافة المشارك {name} بنجاح')

                    # إذا كان الطلب من النافذة الجديدة، أضف معرف خاص
                    if request.POST.get('from_new_modal') == 'true':
                        messages.info(request, 'modal_success_close')

                except ValueError:
                    messages.error(request, 'قيمة الصف غير صالحة')

                return redirect('competitions:start_competition')

            elif add_type == 'multiple':
                # New multiple participants addition
                participants_text = request.POST.get('participants_text', '').strip()
                bulk_grade = request.POST.get('bulk_grade', '').strip()
                bulk_group = request.POST.get('bulk_group', '1').strip()

                if not participants_text:
                    messages.error(request, 'الرجاء إدخال أسماء المشاركين')
                    return redirect('competitions:start_competition')

                if not bulk_grade:
                    messages.error(request, 'الرجاء اختيار الصف للمشاركين')
                    return redirect('competitions:start_competition')

                # Convert grade and group to integer and validate
                try:
                    bulk_grade = int(bulk_grade)
                    if bulk_grade < 1 or bulk_grade > 9:
                        messages.error(request, 'الصف غير صالح')
                        return redirect('competitions:start_competition')

                    bulk_group = int(bulk_group)
                    if bulk_group not in [1, 2]:
                        bulk_group = 1

                    # Parse participant names
                    participant_names = [name.strip() for name in participants_text.split('\n') if name.strip()]

                    if not participant_names:
                        messages.error(request, 'لم يتم العثور على أسماء صالحة')
                        return redirect('competitions:start_competition')

                    # Add participants
                    added_count = 0
                    skipped_count = 0
                    skipped_names = []

                    for name in participant_names:
                        if len(name) > 100:  # Validate name length
                            skipped_count += 1
                            skipped_names.append(f"{name[:50]}... (اسم طويل)")
                            continue

                        # Check if participant already exists
                        if Participant.objects.filter(name=name, grade=bulk_grade, group=bulk_group).exists():
                            skipped_count += 1
                            skipped_names.append(f"{name} (موجود بالفعل)")
                            continue

                        # Create the participant
                        Participant.objects.create(
                            name=name,
                            grade=bulk_grade,
                            group=bulk_group
                        )
                        added_count += 1

                    # Show results
                    if added_count > 0:
                        messages.success(request, f'✅ تمت إضافة {added_count} مشارك بنجاح')

                        # إذا كان الطلب من النافذة الجديدة، أضف معرف خاص
                        if request.POST.get('from_new_modal') == 'true':
                            messages.info(request, 'modal_success_close')

                    if skipped_count > 0:
                        skipped_list = ', '.join(skipped_names[:5])  # Show first 5 skipped names
                        if len(skipped_names) > 5:
                            skipped_list += f' و {len(skipped_names) - 5} آخرين'
                        messages.warning(request, f'⚠️ تم تخطي {skipped_count} مشارك: {skipped_list}')

                except ValueError:
                    messages.error(request, 'قيمة الصف غير صالحة')

                return redirect('competitions:start_competition')

            else:
                # Legacy single participant addition (for backward compatibility)
                name = request.POST.get('participant_name')
                grade = request.POST.get('grade')
                group = request.POST.get('group', '1')

                if not name or not grade:
                    messages.error(request, 'الرجاء ملء جميع الحقول المطلوبة')
                    return redirect('competitions:start_competition')

                # Convert grade and group to integer and validate
                try:
                    grade = int(grade)
                    if grade < 1 or grade > 9:
                        messages.error(request, 'الصف غير صالح')
                        return redirect('competitions:start_competition')

                    # Validate group
                    group = int(group)
                    if group not in [1, 2]:
                        group = 1

                    # Create the participant
                    Participant.objects.create(
                        name=name,
                        grade=grade,
                        group=group
                    )

                    messages.success(request, f'تمت إضافة المشارك {name} بنجاح')
                except ValueError:
                    messages.error(request, 'قيمة الصف غير صالحة')

                return redirect('competitions:start_competition')

    # If not POST, redirect to start page
    return redirect('competitions:start_competition')

@login_required
def delete_participant(request, participant_id):
    """Delete a participant and all their competitions."""
    participant = get_object_or_404(Participant, id=participant_id)
    participant_name = participant.name

    # Count competitions before deletion for the success message
    competitions_count = participant.competitions.count()

    # Delete the participant (this will cascade delete all related competitions due to CASCADE)
    participant.delete()

    # Show success message with details
    if competitions_count > 0:
        messages.success(request, f'✅ تم حذف المشارك {participant_name} وجميع مسابقاته ({competitions_count} مسابقة) بنجاح')
    else:
        messages.success(request, f'✅ تم حذف المشارك {participant_name} بنجاح')

    # Redirect to the competitions start competition page
    return redirect('competitions:start_competition')

@login_required
@require_POST
def clear_competition_history(request):
    """Delete all competition history for the current user"""
    # Get all competitions for this user
    competitions = Competition.objects.filter(user=request.user)

    # Count how many were deleted
    deleted_count = competitions.count()

    if deleted_count > 0:
        # Delete all competitions
        competitions.delete()
        messages.success(request, f'تم حذف {deleted_count} مسابقة بنجاح')
    else:
        messages.info(request, 'لا توجد مسابقات للحذف')

    return redirect('competitions:history')

@login_required
@require_POST
def clear_all_participants(request):
    """Delete all participants and their competitions"""
    # Get all participants
    participants = Participant.objects.all()

    # Count participants and competitions before deletion
    deleted_count = participants.count()
    total_competitions = 0

    for participant in participants:
        total_competitions += participant.competitions.count()

    if deleted_count > 0:
        # Delete all participants (this will cascade delete all competitions)
        participants.delete()

        if total_competitions > 0:
            messages.success(request, f'✅ تم حذف جميع المشاركين ({deleted_count} مشارك) وجميع مسابقاتهم ({total_competitions} مسابقة) بنجاح')
        else:
            messages.success(request, f'✅ تم حذف جميع المشاركين ({deleted_count} مشارك) بنجاح')
    else:
        messages.info(request, 'لا توجد مشاركين للحذف')

    return redirect('competitions:start_competition')

@login_required
def delete_multiple_participants(request):
    """Delete multiple participants and all their competitions at once."""
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    # Get the list of participant IDs from the form
    participant_ids = request.POST.getlist('participant_ids')

    if not participant_ids:
        messages.error(request, 'لم يتم تحديد أي مشارك للحذف')
        return redirect('competitions:start_competition')

    # Count participants and their competitions before deletion
    deleted_count = 0
    total_competitions_deleted = 0
    participant_names = []

    for participant_id in participant_ids:
        try:
            participant = Participant.objects.get(id=participant_id)
            participant_names.append(participant.name)

            # Count competitions before deletion
            competitions_count = participant.competitions.count()
            total_competitions_deleted += competitions_count

            # Delete the participant (this will cascade delete all related competitions)
            participant.delete()
            deleted_count += 1

        except Participant.DoesNotExist:
            continue

    # Show detailed success message
    if deleted_count > 0:
        if total_competitions_deleted > 0:
            messages.success(request, f'✅ تم حذف {deleted_count} مشارك وجميع مسابقاتهم ({total_competitions_deleted} مسابقة) بنجاح')
        else:
            messages.success(request, f'✅ تم حذف {deleted_count} مشارك بنجاح')
    else:
        messages.warning(request, 'لم يتم حذف أي مشارك')

    # Redirect to the competitions start competition page
    return redirect('competitions:start_competition')

@login_required
def get_participants_by_grade(request):
    """Get participants by grade via AJAX"""
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        grade = request.GET.get('grade')
        if grade:
            participants = Participant.objects.filter(grade=grade).order_by('name')
            participants_data = []

            for participant in participants:
                participants_data.append({
                    'id': participant.id,
                    'name': participant.name,
                    'grade': participant.grade,
                    'group': participant.group,
                    'grade_display': participant.get_grade_display(),
                    'group_display': participant.get_group_display(),
                    'initials': participant.name[:2].upper() if len(participant.name) >= 2 else participant.name[:1].upper(),
                    'short_name': f"{participant.name[:1].upper()}{participant.name[1:2].upper()}." if len(participant.name) >= 2 else participant.name,
                    'short_grade': participant.get_grade_display()[:6],
                    'short_group': participant.get_group_display()[:6]
                })

            return JsonResponse({
                'success': True,
                'participants': participants_data,
                'count': len(participants_data)
            })
        else:
            return JsonResponse({'success': False, 'error': 'Grade parameter is required'})

    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def score_test(request):
    """صفحة اختبار عرض النتائج بالصيغ المختلفة"""
    return render(request, 'competitions/score_test.html')

@login_required
def get_chart_data(request, chart_type):
    """API endpoint to get chart data for analytics dashboard"""
    try:
        if chart_type == 'grade_stats':
            # Get grade-level statistics
            grade_stats = []
            for grade_value, grade_display in Participant.GRADE_CHOICES:
                participants = Participant.objects.filter(grade=grade_value)
                competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
                responses = UserResponse.objects.filter(competition__in=competitions)

                total_participants = participants.count()
                total_competitions = competitions.count()
                total_correct = responses.filter(is_correct=True).count()
                total_questions = responses.count()

                avg_response_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
                success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0

                if total_participants > 0:  # Only include grades with participants
                    grade_stats.append({
                        'grade': grade_display,
                        'participants': total_participants,
                        'competitions': total_competitions,
                        'successRate': round(success_rate, 2),
                        'avgTime': round(avg_response_time, 2),
                        'totalQuestions': total_questions
                    })

            return JsonResponse({'data': grade_stats})

        elif chart_type == 'operations_stats':
            # Get operations statistics
            operations_data = []
            operation_names = {
                'addition': 'الجمع',
                'subtraction': 'الطرح',
                'multiplication': 'الضرب',
                'division': 'القسمة'
            }

            for operation, operation_name in operation_names.items():
                responses = UserResponse.objects.filter(question__operation=operation)
                total_questions = responses.count()
                correct_answers = responses.filter(is_correct=True).count()

                success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0
                avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

                operations_data.append({
                    'operation': operation_name,
                    'questionsCount': total_questions,
                    'successRate': round(success_rate, 2),
                    'avgTime': round(avg_time, 2)
                })

            return JsonResponse({'data': operations_data})

        elif chart_type == 'time_trends':
            # Get time-based trends (last 30 days)
            from datetime import datetime, timedelta
            from django.db.models.functions import TruncDate
            thirty_days_ago = timezone.now() - timedelta(days=30)

            # Group competitions by day
            daily_stats = Competition.objects.filter(
                is_completed=True,
                end_time__gte=thirty_days_ago
            ).annotate(
                day=TruncDate('end_time')
            ).values('day').annotate(
                count=Count('id'),
                avg_score=Avg('result__total_score')
            ).order_by('day')

            trends_data = []
            for stat in daily_stats:
                trends_data.append({
                    'date': stat['day'].strftime('%Y-%m-%d') if stat['day'] else '',
                    'competitions': stat['count'],
                    'avgScore': round(stat['avg_score'] or 0, 2)
                })

            return JsonResponse({'data': trends_data})

        else:
            return JsonResponse({'error': 'نوع الرسم البياني غير مدعوم'}, status=400)

    except Exception as e:
        return JsonResponse({'error': f'خطأ في جلب البيانات: {str(e)}'}, status=500)

@login_required
def export_grade_analytics_excel(request):
    """Export grade-level analytics as Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime

        # Create a BytesIO buffer to receive the Excel file
        buffer = io.BytesIO()

        # Get grade statistics
        grade_data = []
        for grade_value, grade_display in Participant.GRADE_CHOICES:
            participants = Participant.objects.filter(grade=grade_value)
            competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
            responses = UserResponse.objects.filter(competition__in=competitions)

            total_participants = participants.count()
            total_competitions = competitions.count()
            total_correct = responses.filter(is_correct=True).count()
            total_questions = responses.count()

            avg_response_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
            success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0

            if total_participants > 0:  # Only include grades with participants
                grade_data.append({
                    'المستوى الدراسي': grade_display,
                    'عدد المشاركين': total_participants,
                    'إجمالي المسابقات': total_competitions,
                    'متوسط وقت الإجابة (ثانية)': round(avg_response_time, 2),
                    'معدل النجاح (%)': round(success_rate, 2),
                    'إجمالي الأسئلة': total_questions,
                    'الإجابات الصحيحة': total_correct,
                    'الإجابات الخاطئة': total_questions - total_correct
                })

        # Create Excel file with multiple sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main grade statistics
            if grade_data:
                grade_df = pd.DataFrame(grade_data)
                grade_df.to_excel(writer, sheet_name='إحصائيات المستويات', index=False)

                # Summary sheet
                summary_data = [{
                    'إجمالي المستويات': len(grade_data),
                    'إجمالي المشاركين': sum(item['عدد المشاركين'] for item in grade_data),
                    'إجمالي المسابقات': sum(item['إجمالي المسابقات'] for item in grade_data),
                    'متوسط النجاح العام (%)': round(sum(item['معدل النجاح (%)'] for item in grade_data) / len(grade_data), 2) if grade_data else 0,
                    'تاريخ التصدير': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='الملخص العام', index=False)
            else:
                # Empty data message
                empty_df = pd.DataFrame([{'رسالة': 'لا توجد بيانات للتصدير'}])
                empty_df.to_excel(writer, sheet_name='لا توجد بيانات', index=False)

        # Prepare response
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Generate filename with current date (using ASCII-safe filename)
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f'Grade_Analytics_{date_str}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'تحليلات_المستويات_الدراسية_{date_str}.xlsx'

        return response

    except Exception as e:
        # Return error as Excel file
        error_buffer = io.BytesIO()
        error_data = [{
            'خطأ': f'فشل في تصدير البيانات: {str(e)}',
            'تاريخ الخطأ': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }]
        error_df = pd.DataFrame(error_data)
        error_df.to_excel(error_buffer, index=False, sheet_name='خطأ')
        error_buffer.seek(0)

        response = HttpResponse(error_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_تصدير_المستويات.xlsx"'
        return response

@login_required
def export_operations_analytics_excel(request):
    """Export operations analytics as Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime

        # Create a BytesIO buffer to receive the Excel file
        buffer = io.BytesIO()

        # Get operations statistics
        operations_data = []
        operation_names = {
            'addition': 'الجمع',
            'subtraction': 'الطرح',
            'multiplication': 'الضرب',
            'division': 'القسمة'
        }

        for operation, operation_name in operation_names.items():
            responses = UserResponse.objects.filter(question__operation=operation)
            total_questions = responses.count()
            correct_answers = responses.filter(is_correct=True).count()

            success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0
            avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

            # Get strengths and weaknesses
            strengths, weaknesses = analyze_operation_patterns(operation, responses)

            operations_data.append({
                'العملية الحسابية': operation_name,
                'عدد الأسئلة': total_questions,
                'الإجابات الصحيحة': correct_answers,
                'الإجابات الخاطئة': total_questions - correct_answers,
                'معدل النجاح (%)': round(success_rate, 2),
                'متوسط وقت الإجابة (ثانية)': round(avg_time, 2),
                'نقاط القوة': ', '.join(strengths) if strengths else 'لا توجد',
                'نقاط الضعف': ', '.join(weaknesses) if weaknesses else 'لا توجد'
            })

        # Create Excel file with multiple sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main operations statistics
            if operations_data:
                operations_df = pd.DataFrame(operations_data)
                operations_df.to_excel(writer, sheet_name='تحليل العمليات', index=False)

                # Detailed analysis by grade level
                grade_operations_data = []
                for grade_value, grade_display in Participant.GRADE_CHOICES:
                    participants = Participant.objects.filter(grade=grade_value)
                    if participants.exists():
                        for operation, operation_name in operation_names.items():
                            responses = UserResponse.objects.filter(
                                competition__participant__in=participants,
                                question__operation=operation
                            )
                            total_questions = responses.count()
                            correct_answers = responses.filter(is_correct=True).count()
                            success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0
                            avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

                            if total_questions > 0:
                                grade_operations_data.append({
                                    'المستوى الدراسي': grade_display,
                                    'العملية': operation_name,
                                    'عدد الأسئلة': total_questions,
                                    'معدل النجاح (%)': round(success_rate, 2),
                                    'متوسط الوقت (ثانية)': round(avg_time, 2)
                                })

                if grade_operations_data:
                    grade_operations_df = pd.DataFrame(grade_operations_data)
                    grade_operations_df.to_excel(writer, sheet_name='العمليات حسب المستوى', index=False)

                # Summary sheet
                summary_data = [{
                    'إجمالي العمليات': len(operations_data),
                    'إجمالي الأسئلة': sum(item['عدد الأسئلة'] for item in operations_data),
                    'إجمالي الإجابات الصحيحة': sum(item['الإجابات الصحيحة'] for item in operations_data),
                    'متوسط النجاح العام (%)': round(sum(item['معدل النجاح (%)'] for item in operations_data) / len(operations_data), 2) if operations_data else 0,
                    'تاريخ التصدير': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='الملخص العام', index=False)
            else:
                # Empty data message
                empty_df = pd.DataFrame([{'رسالة': 'لا توجد بيانات للتصدير'}])
                empty_df.to_excel(writer, sheet_name='لا توجد بيانات', index=False)

        # Prepare response
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Generate filename with current date (using ASCII-safe filename)
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f'Operations_Analytics_{date_str}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'تحليلات_العمليات_الحسابية_{date_str}.xlsx'

        return response

    except Exception as e:
        # Return error as Excel file
        error_buffer = io.BytesIO()
        error_data = [{
            'خطأ': f'فشل في تصدير البيانات: {str(e)}',
            'تاريخ الخطأ': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }]
        error_df = pd.DataFrame(error_data)
        error_df.to_excel(error_buffer, index=False, sheet_name='خطأ')
        error_buffer.seek(0)

        response = HttpResponse(error_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_تصدير_العمليات.xlsx"'
        return response

@login_required
def export_general_analytics_excel(request):
    """Export comprehensive general analytics as Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime, timedelta

        # Create a BytesIO buffer to receive the Excel file
        buffer = io.BytesIO()

        # Create Excel file with multiple comprehensive sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

            # 1. Overall System Summary
            total_participants = Participant.objects.count()
            total_competitions = Competition.objects.filter(is_completed=True).count()
            total_questions = UserResponse.objects.count()
            total_correct = UserResponse.objects.filter(is_correct=True).count()
            overall_success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0
            avg_response_time = UserResponse.objects.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

            system_summary = [{
                'إجمالي المشاركين': total_participants,
                'إجمالي المسابقات المكتملة': total_competitions,
                'إجمالي الأسئلة': total_questions,
                'إجمالي الإجابات الصحيحة': total_correct,
                'إجمالي الإجابات الخاطئة': total_questions - total_correct,
                'معدل النجاح العام (%)': round(overall_success_rate, 2),
                'متوسط وقت الإجابة العام (ثانية)': round(avg_response_time, 2),
                'تاريخ التقرير': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }]
            summary_df = pd.DataFrame(system_summary)
            summary_df.to_excel(writer, sheet_name='الملخص العام للنظام', index=False)

            # 2. Grade Level Statistics
            grade_data = []
            for grade_value, grade_display in Participant.GRADE_CHOICES:
                participants = Participant.objects.filter(grade=grade_value)
                competitions = Competition.objects.filter(participant__in=participants, is_completed=True)
                responses = UserResponse.objects.filter(competition__in=competitions)

                total_participants_grade = participants.count()
                total_competitions_grade = competitions.count()
                total_correct_grade = responses.filter(is_correct=True).count()
                total_questions_grade = responses.count()

                avg_response_time_grade = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
                success_rate_grade = (total_correct_grade / total_questions_grade * 100) if total_questions_grade > 0 else 0

                if total_participants_grade > 0:
                    grade_data.append({
                        'المستوى الدراسي': grade_display,
                        'عدد المشاركين': total_participants_grade,
                        'إجمالي المسابقات': total_competitions_grade,
                        'إجمالي الأسئلة': total_questions_grade,
                        'الإجابات الصحيحة': total_correct_grade,
                        'الإجابات الخاطئة': total_questions_grade - total_correct_grade,
                        'معدل النجاح (%)': round(success_rate_grade, 2),
                        'متوسط وقت الإجابة (ثانية)': round(avg_response_time_grade, 2),
                        'متوسط المسابقات لكل مشارك': round(total_competitions_grade / total_participants_grade, 2) if total_participants_grade > 0 else 0
                    })

            if grade_data:
                grade_df = pd.DataFrame(grade_data)
                grade_df.to_excel(writer, sheet_name='إحصائيات المستويات الدراسية', index=False)

            # 3. Operations Performance Analysis
            operations_data = []
            operation_names = {
                'addition': 'الجمع',
                'subtraction': 'الطرح',
                'multiplication': 'الضرب',
                'division': 'القسمة'
            }

            for operation, operation_name in operation_names.items():
                responses = UserResponse.objects.filter(question__operation=operation)
                total_questions_op = responses.count()
                correct_answers_op = responses.filter(is_correct=True).count()
                success_rate_op = (correct_answers_op / total_questions_op * 100) if total_questions_op > 0 else 0
                avg_time_op = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

                operations_data.append({
                    'العملية الحسابية': operation_name,
                    'عدد الأسئلة': total_questions_op,
                    'الإجابات الصحيحة': correct_answers_op,
                    'الإجابات الخاطئة': total_questions_op - correct_answers_op,
                    'معدل النجاح (%)': round(success_rate_op, 2),
                    'متوسط وقت الإجابة (ثانية)': round(avg_time_op, 2),
                    'نسبة من إجمالي الأسئلة (%)': round((total_questions_op / total_questions * 100), 2) if total_questions > 0 else 0
                })

            if operations_data:
                operations_df = pd.DataFrame(operations_data)
                operations_df.to_excel(writer, sheet_name='تحليل العمليات الحسابية', index=False)

            # 4. Top Performers
            top_performers_data = []
            top_by_accuracy = Participant.objects.annotate(
                total_responses=Count('competitions__responses'),
                correct_responses=Count('competitions__responses', filter=Q(competitions__responses__is_correct=True)),
                accuracy=Case(
                    When(total_responses=0, then=0.0),
                    default=F('correct_responses') * 100.0 / F('total_responses'),
                    output_field=IntegerField()
                )
            ).filter(total_responses__gte=5).order_by('-accuracy')[:10]

            for participant in top_by_accuracy:
                avg_time = UserResponse.objects.filter(
                    competition__participant=participant
                ).aggregate(avg_time=Avg('response_time'))['avg_time'] or 0

                top_performers_data.append({
                    'اسم المشارك': participant.name,
                    'المستوى الدراسي': participant.get_grade_display(),
                    'إجمالي الإجابات': participant.total_responses,
                    'الإجابات الصحيحة': participant.correct_responses,
                    'معدل الدقة (%)': round(participant.accuracy, 2),
                    'متوسط وقت الإجابة (ثانية)': round(avg_time, 2),
                    'عدد المسابقات': participant.competitions.filter(is_completed=True).count()
                })

            if top_performers_data:
                top_performers_df = pd.DataFrame(top_performers_data)
                top_performers_df.to_excel(writer, sheet_name='أفضل المؤدين', index=False)

            # 5. Time-based Trends (Last 30 days)
            thirty_days_ago = timezone.now() - timedelta(days=30)
            from django.db.models.functions import TruncDate
            daily_stats = Competition.objects.filter(
                is_completed=True,
                end_time__gte=thirty_days_ago
            ).annotate(
                day=TruncDate('end_time')
            ).values('day').annotate(
                competitions_count=Count('id'),
                avg_score=Avg('result__total_score'),
                total_questions=Count('responses'),
                correct_answers=Count('responses', filter=Q(responses__is_correct=True))
            ).order_by('day')

            trends_data = []
            for stat in daily_stats:
                success_rate_daily = (stat['correct_answers'] / stat['total_questions'] * 100) if stat['total_questions'] > 0 else 0
                trends_data.append({
                    'التاريخ': stat['day'].strftime('%Y-%m-%d') if stat['day'] else '',
                    'عدد المسابقات': stat['competitions_count'],
                    'متوسط النتيجة': round(stat['avg_score'] or 0, 2),
                    'إجمالي الأسئلة': stat['total_questions'],
                    'الإجابات الصحيحة': stat['correct_answers'],
                    'معدل النجاح اليومي (%)': round(success_rate_daily, 2)
                })

            if trends_data:
                trends_df = pd.DataFrame(trends_data)
                trends_df.to_excel(writer, sheet_name='الاتجاهات الزمنية (30 يوم)', index=False)

            # 6. Difficulty Level Analysis
            difficulty_data = []
            for difficulty in range(1, 10):  # Difficulty levels 1-9
                competitions_diff = Competition.objects.filter(difficulty=difficulty, is_completed=True)
                responses_diff = UserResponse.objects.filter(competition__in=competitions_diff)

                total_competitions_diff = competitions_diff.count()
                total_questions_diff = responses_diff.count()
                correct_answers_diff = responses_diff.filter(is_correct=True).count()
                avg_time_diff = responses_diff.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
                success_rate_diff = (correct_answers_diff / total_questions_diff * 100) if total_questions_diff > 0 else 0

                if total_competitions_diff > 0:
                    difficulty_data.append({
                        'مستوى الصعوبة': f'مستوى {difficulty}',
                        'عدد المسابقات': total_competitions_diff,
                        'إجمالي الأسئلة': total_questions_diff,
                        'الإجابات الصحيحة': correct_answers_diff,
                        'معدل النجاح (%)': round(success_rate_diff, 2),
                        'متوسط وقت الإجابة (ثانية)': round(avg_time_diff, 2),
                        'نسبة الاستخدام (%)': round((total_competitions_diff / total_competitions * 100), 2) if total_competitions > 0 else 0
                    })

            if difficulty_data:
                difficulty_df = pd.DataFrame(difficulty_data)
                difficulty_df.to_excel(writer, sheet_name='تحليل مستويات الصعوبة', index=False)

        # Prepare response
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Generate filename with current date (using ASCII-safe filename)
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f'General_Analytics_{date_str}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'التحليلات_العامة_الشاملة_{date_str}.xlsx'

        return response

    except Exception as e:
        # Return error as Excel file
        error_buffer = io.BytesIO()
        error_data = [{
            'خطأ': f'فشل في تصدير التحليلات العامة: {str(e)}',
            'تاريخ الخطأ': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }]
        error_df = pd.DataFrame(error_data)
        error_df.to_excel(error_buffer, index=False, sheet_name='خطأ')
        error_buffer.seek(0)

        response = HttpResponse(error_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_التحليلات_العامة.xlsx"'
        return response

@login_required
def export_participants_results_excel(request):
    """Export participants results by grade level as Excel file"""
    try:
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime

        # Create a BytesIO buffer to receive the Excel file
        buffer = io.BytesIO()

        # Create Excel file with multiple sheets (one per grade)
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

            # Overall summary sheet
            summary_data = []
            total_participants = 0
            total_competitions = 0

            # Process each grade level
            for grade_value, grade_display in Participant.GRADE_CHOICES:
                participants = Participant.objects.filter(grade=grade_value)

                if participants.exists():
                    grade_data = []
                    grade_participants = 0
                    grade_competitions = 0

                    for participant in participants:
                        # Calculate statistics manually to avoid property issues
                        competitions = participant.competitions.filter(is_completed=True)
                        competitions_count = competitions.count()

                        if competitions_count > 0:
                            # Calculate average score manually
                            total_score = 0
                            best_score = 0
                            total_questions = 0
                            correct_answers = 0
                            avg_response_time = 0

                            for comp in competitions:
                                if hasattr(comp, 'result'):
                                    score = comp.result.total_score
                                    total_score += score
                                    if score > best_score:
                                        best_score = score

                                # Calculate response statistics
                                responses = comp.responses.all()
                                total_questions += responses.count()
                                correct_answers += responses.filter(is_correct=True).count()

                                # Calculate average response time
                                response_times = [r.response_time for r in responses if r.response_time]
                                if response_times:
                                    avg_response_time += sum(response_times) / len(response_times)

                            average_score = total_score / competitions_count if competitions_count > 0 else 0
                            avg_response_time = avg_response_time / competitions_count if competitions_count > 0 else 0
                            success_rate = (correct_answers / total_questions * 100) if total_questions > 0 else 0

                            grade_data.append({
                                'اسم المشارك': participant.name,
                                'الفوج': participant.get_group_display(),
                                'عدد المسابقات': competitions_count,
                                'متوسط النقاط': round(average_score, 2),
                                'أفضل نتيجة': best_score,
                                'إجمالي الأسئلة': total_questions,
                                'الإجابات الصحيحة': correct_answers,
                                'معدل النجاح (%)': round(success_rate, 2),
                                'متوسط وقت الإجابة (ثانية)': round(avg_response_time, 2),
                                'تاريخ آخر مسابقة': competitions.last().end_time.strftime('%Y-%m-%d %H:%M') if competitions.last() and competitions.last().end_time else 'غير مكتملة'
                            })

                            grade_participants += 1
                            grade_competitions += competitions_count

                    # Create sheet for this grade if there's data
                    if grade_data:
                        grade_df = pd.DataFrame(grade_data)
                        sheet_name = f'{grade_display}'
                        # Ensure sheet name is not too long
                        if len(sheet_name) > 31:
                            sheet_name = f'المستوى {grade_value}'
                        grade_df.to_excel(writer, sheet_name=sheet_name, index=False)

                        # Add to summary
                        summary_data.append({
                            'المستوى الدراسي': grade_display,
                            'عدد المشاركين': grade_participants,
                            'إجمالي المسابقات': grade_competitions,
                            'متوسط المسابقات لكل مشارك': round(grade_competitions / grade_participants, 2) if grade_participants > 0 else 0
                        })

                        total_participants += grade_participants
                        total_competitions += grade_competitions

            # Create summary sheet
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='الملخص العام', index=False)

                # Add overall statistics
                overall_stats = [{
                    'إجمالي المشاركين النشطين': total_participants,
                    'إجمالي المسابقات': total_competitions,
                    'متوسط المسابقات لكل مشارك': round(total_competitions / total_participants, 2) if total_participants > 0 else 0,
                    'عدد المستويات النشطة': len(summary_data),
                    'تاريخ التقرير': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }]
                overall_df = pd.DataFrame(overall_stats)
                overall_df.to_excel(writer, sheet_name='الإحصائيات العامة', index=False)
            else:
                # No data message
                empty_df = pd.DataFrame([{'رسالة': 'لا توجد بيانات مشاركين للتصدير'}])
                empty_df.to_excel(writer, sheet_name='لا توجد بيانات', index=False)

        # Prepare response
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Generate filename with current date
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f'Participants_Results_{date_str}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'نتائج_المشاركين_{date_str}.xlsx'

        return response

    except Exception as e:
        # Return error as Excel file
        error_buffer = io.BytesIO()
        error_data = [{
            'خطأ': f'فشل في تصدير نتائج المشاركين: {str(e)}',
            'تاريخ الخطأ': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'تفاصيل الخطأ': str(type(e).__name__)
        }]
        error_df = pd.DataFrame(error_data)
        error_df.to_excel(error_buffer, index=False, sheet_name='خطأ')
        error_buffer.seek(0)

        response = HttpResponse(error_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="خطأ_في_تصدير_نتائج_المشاركين.xlsx"'
        return response

def analyze_operation_patterns(operation, responses):
    """Analyze strengths and weaknesses for a specific operation"""
    strengths = []
    weaknesses = []

    if not responses.exists():
        return strengths, weaknesses

    # Analyze response times
    avg_time = responses.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
    fast_responses = responses.filter(response_time__lt=avg_time * 0.8).count()
    slow_responses = responses.filter(response_time__gt=avg_time * 1.5).count()

    if fast_responses > slow_responses:
        strengths.append('سرعة في الإجابة')
    elif slow_responses > fast_responses:
        weaknesses.append('بطء في الإجابة')

    # Analyze accuracy
    total_responses = responses.count()
    correct_responses = responses.filter(is_correct=True).count()
    accuracy_rate = (correct_responses / total_responses * 100) if total_responses > 0 else 0

    if accuracy_rate >= 80:
        strengths.append('دقة عالية في الإجابات')
    elif accuracy_rate >= 60:
        strengths.append('دقة متوسطة في الإجابات')
    else:
        weaknesses.append('دقة منخفضة في الإجابات')

    # Analyze difficulty patterns
    easy_questions = responses.filter(question__difficulty__lte=3)
    hard_questions = responses.filter(question__difficulty__gte=7)

    if easy_questions.exists():
        easy_accuracy = easy_questions.filter(is_correct=True).count() / easy_questions.count() * 100
        if easy_accuracy >= 90:
            strengths.append('إتقان الأسئلة السهلة')
        elif easy_accuracy < 70:
            weaknesses.append('صعوبة في الأسئلة السهلة')

    if hard_questions.exists():
        hard_accuracy = hard_questions.filter(is_correct=True).count() / hard_questions.count() * 100
        if hard_accuracy >= 60:
            strengths.append('قدرة على حل الأسئلة الصعبة')
        elif hard_accuracy < 30:
            weaknesses.append('صعوبة في الأسئلة المعقدة')

    return strengths, weaknesses


# ==================== Student Views ====================

def home_selection(request):
    """صفحة اختيار نوع الدخول (أستاذ أو تلميذ)"""
    return render(request, 'home_selection.html')


def student_login(request):
    """صفحة دخول التلميذ برمز"""
    if request.method == 'POST':
        access_code = request.POST.get('access_code', '').strip().lower()

        # التحقق من رمز الدخول
        if access_code == 'ben25':
            # حفظ رمز الدخول في الجلسة
            request.session['student_access_code'] = access_code
            request.session['student_authenticated'] = True
            messages.success(request, 'تم التحقق من رمز الدخول بنجاح! 🎉')
            return redirect('student:setup')
        else:
            messages.error(request, 'رمز الدخول غير صحيح. يرجى المحاولة مرة أخرى.')

    return render(request, 'student/login.html')


def student_setup(request):
    """صفحة إعداد التلميذ (الاسم والمستوى والصعوبة)"""
    # التحقق من تسجيل الدخول
    if not request.session.get('student_authenticated'):
        messages.error(request, 'يرجى إدخال رمز الدخول أولاً')
        return redirect('student:login')

    if request.method == 'POST':
        student_name = request.POST.get('student_name', '').strip()
        grade = request.POST.get('grade')
        difficulty = request.POST.get('difficulty')

        # التحقق من البيانات
        if not student_name or not grade or not difficulty:
            messages.error(request, 'يرجى إكمال جميع البيانات المطلوبة')
            return render(request, 'student/setup.html')

        try:
            grade = int(grade)
            difficulty = int(difficulty)

            # التحقق من صحة القيم
            if grade < 1 or grade > 9:
                messages.error(request, 'المستوى الدراسي غير صالح')
                return render(request, 'student/setup.html')

            if difficulty < 1 or difficulty > 9:
                messages.error(request, 'مستوى الصعوبة غير صالح')
                return render(request, 'student/setup.html')

            # إنشاء جلسة جديدة للتلميذ
            session = StudentSession.objects.create(
                student_name=student_name,
                grade=grade,
                difficulty=difficulty,
                session_code=request.session.get('student_access_code', 'ben25'),
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )

            # حفظ معرف الجلسة
            request.session['student_session_id'] = session.id

            messages.success(request, f'مرحباً {student_name}! جاري تحضير مسابقتك... 🚀')
            return redirect('student:competition')

        except ValueError:
            messages.error(request, 'بيانات غير صالحة')

    return render(request, 'student/setup.html')


def student_competition(request):
    """صفحة المسابقة للتلميذ"""
    # التحقق من الجلسة
    session_id = request.session.get('student_session_id')
    if not session_id:
        messages.error(request, 'جلسة غير صالحة. يرجى البدء من جديد.')
        return redirect('student:login')

    try:
        session = StudentSession.objects.get(id=session_id, is_completed=False)
    except StudentSession.DoesNotExist:
        messages.error(request, 'جلسة غير موجودة أو مكتملة')
        return redirect('student:login')

    # إذا كانت المسابقة مكتملة، اذهب للنتائج
    if session.is_completed:
        return redirect('student:results')

    context = {
        'session': session,
        'student_name': session.student_name,
        'grade_display': session.get_grade_display(),
        'difficulty_display': session.get_difficulty_display(),
    }

    return render(request, 'student/competition.html', context)


@csrf_exempt
def student_get_question(request):
    """API لجلب سؤال جديد للتلميذ"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    session_id = request.session.get('student_session_id')
    if not session_id:
        return JsonResponse({'error': 'Invalid session'}, status=400)

    try:
        session = StudentSession.objects.get(id=session_id, is_completed=False)
    except StudentSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)

    # توليد سؤال جديد باستخدام نفس النظام المستخدم في المنصة الأصلية
    operations = ['addition', 'subtraction', 'multiplication', 'division']
    operation = random.choice(operations)

    try:
        # للمستويات المتدرجة (7-9)، استخدام نظام خاص
        if session.difficulty in [7, 8, 9]:
            result = generate_progressive_question(session)
            question_data = result['question_data']
            operation = result['operation']
        else:
            # استخدام QuestionGenerator مثل النظام الأصلي للمستويات 1-6
            question_data = generate_math_question_new(operation, session.difficulty)

        return JsonResponse({
            'success': True,
            'question': {
                'operation': operation,
                'first_number': question_data['first_number'],
                'second_number': question_data['second_number'],
                'correct_answer': question_data['answer'],
                'operation_symbol': get_operation_symbol(operation)
            }
        })
    except Exception as e:
        return JsonResponse({'error': f'Error generating question: {str(e)}'}, status=500)


@csrf_exempt
def student_submit_answer(request):
    """API لإرسال إجابة التلميذ"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    session_id = request.session.get('student_session_id')
    if not session_id:
        return JsonResponse({'error': 'Invalid session'}, status=400)

    try:
        session = StudentSession.objects.get(id=session_id, is_completed=False)
    except StudentSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)

    try:
        data = json.loads(request.body)

        operation = data.get('operation')
        first_number = int(data.get('first_number'))
        second_number = int(data.get('second_number'))
        correct_answer = int(data.get('correct_answer'))
        student_answer = data.get('student_answer')
        response_time = float(data.get('response_time', 0))

        # التحقق من الإجابة
        is_correct = False
        if student_answer is not None:
            try:
                student_answer = int(student_answer)
                is_correct = student_answer == correct_answer
            except (ValueError, TypeError):
                student_answer = None

        # حفظ الإجابة
        response = StudentResponse.objects.create(
            session=session,
            operation=operation,
            first_number=first_number,
            second_number=second_number,
            correct_answer=correct_answer,
            student_answer=student_answer,
            is_correct=is_correct,
            response_time=response_time
        )

        # تحديث إحصائيات الجلسة
        session.total_questions += 1
        if is_correct:
            session.correct_answers += 1
            session.total_score += 3  # 3 نقاط لكل إجابة صحيحة

        # تحديث إحصائيات العمليات
        if operation == 'addition':
            session.addition_total += 1
            if is_correct:
                session.addition_correct += 1
        elif operation == 'subtraction':
            session.subtraction_total += 1
            if is_correct:
                session.subtraction_correct += 1
        elif operation == 'multiplication':
            session.multiplication_total += 1
            if is_correct:
                session.multiplication_correct += 1
        elif operation == 'division':
            session.division_total += 1
            if is_correct:
                session.division_correct += 1

        session.save()

        return JsonResponse({
            'success': True,
            'is_correct': is_correct,
            'correct_answer': correct_answer,
            'total_questions': session.total_questions,
            'correct_answers': session.correct_answers,
            'total_score': session.total_score
        })

    except Exception as e:
        return JsonResponse({'error': f'Error processing answer: {str(e)}'}, status=500)


def student_finish_competition(request):
    """إنهاء المسابقة للتلميذ"""
    session_id = request.session.get('student_session_id')
    if not session_id:
        messages.error(request, 'جلسة غير صالحة')
        return redirect('student_login')

    try:
        session = StudentSession.objects.get(id=session_id)
        session.is_completed = True
        session.end_time = timezone.now()
        session.save()

        return redirect('student:results')
    except StudentSession.DoesNotExist:
        messages.error(request, 'جلسة غير موجودة')
        return redirect('student:login')


def student_results(request):
    """صفحة النتائج للتلميذ"""
    session_id = request.session.get('student_session_id')
    if not session_id:
        messages.error(request, 'جلسة غير صالحة')
        return redirect('student:login')

    try:
        session = StudentSession.objects.get(id=session_id, is_completed=True)
    except StudentSession.DoesNotExist:
        messages.error(request, 'جلسة غير موجودة أو غير مكتملة')
        return redirect('student:login')

    context = {
        'session': session,
        'student_name': session.student_name,
        'grade_display': session.get_grade_display(),
        'difficulty_display': session.get_difficulty_display(),
        'duration_minutes': session.duration_minutes,
        'success_rate': session.success_rate,
        'strengths': session.strengths,
        'weaknesses': session.weaknesses,
    }

    return render(request, 'student/results.html', context)


def generate_progressive_question(session):
    """توليد سؤال متدرج الصعوبة حسب عدد الأسئلة المجابة"""
    question_count = session.total_questions

    if session.difficulty == 7:  # الجمع والطرح
        # تحديد العملية (جمع أو طرح)
        operation = 'addition' if random.choice([True, False]) else 'subtraction'

        # تدرج الصعوبة حسب عدد الأسئلة
        if question_count < 5:
            # سهل: أرقام 1-20
            min_val, max_val = 1, 20
        elif question_count < 10:
            # متوسط: أرقام 10-50
            min_val, max_val = 10, 50
        elif question_count < 15:
            # صعب: أرقام 20-100
            min_val, max_val = 20, 100
        else:
            # صعب جداً: أرقام 50-200
            min_val, max_val = 50, 200

    elif session.difficulty == 8:  # الضرب
        operation = 'multiplication'

        if question_count < 5:
            # سهل: جدول الضرب الأساسي
            min_val, max_val = 1, 10
        elif question_count < 10:
            # متوسط: أرقام أكبر
            min_val, max_val = 2, 15
        elif question_count < 15:
            # صعب: أرقام كبيرة
            min_val, max_val = 5, 25
        else:
            # صعب جداً
            min_val, max_val = 10, 50

    elif session.difficulty == 9:  # القسمة
        operation = 'division'

        if question_count < 5:
            # سهل: قسمة بسيطة
            quotients = [2, 3, 4, 5]
            divisors = range(2, 10)
        elif question_count < 10:
            # متوسط
            quotients = [2, 3, 4, 5, 6, 7, 8, 9]
            divisors = range(2, 15)
        elif question_count < 15:
            # صعب
            quotients = range(2, 15)
            divisors = range(2, 20)
        else:
            # صعب جداً
            quotients = range(2, 25)
            divisors = range(2, 30)

    # توليد السؤال باستخدام QuestionGenerator
    from .question_generators import QuestionGenerator

    if operation == 'addition':
        question_data = generate_custom_addition(min_val, max_val)
    elif operation == 'subtraction':
        question_data = generate_custom_subtraction(min_val, max_val)
    elif operation == 'multiplication':
        question_data = generate_custom_multiplication(min_val, max_val)
    elif operation == 'division':
        question_data = generate_custom_division(quotients, divisors)

    return {
        'operation': operation,
        'question_data': question_data
    }


def generate_custom_addition(min_val, max_val):
    """توليد سؤال جمع مخصص"""
    for _ in range(50):
        num1 = random.randint(min_val, max_val)
        num2 = random.randint(min_val, max_val)

        # تجنب الرقم 10
        if num1 == 10 or num2 == 10:
            continue

        first_number = min(num1, num2)
        second_number = max(num1, num2)
        answer = first_number + second_number

        # تجنب الرقم 10 في الناتج
        if '10' not in str(answer):
            return {
                'first_number': first_number,
                'second_number': second_number,
                'answer': answer
            }

    # fallback
    return {'first_number': 5, 'second_number': 7, 'answer': 12}


def generate_custom_subtraction(min_val, max_val):
    """توليد سؤال طرح مخصص"""
    for _ in range(50):
        num1 = random.randint(min_val, max_val)
        num2 = random.randint(min_val, num1)  # للتأكد من أن الناتج موجب

        if num1 == 10 or num2 == 10:
            continue

        first_number = num1
        second_number = num2
        answer = first_number - second_number

        if answer > 0 and '10' not in str(answer):
            return {
                'first_number': first_number,
                'second_number': second_number,
                'answer': answer
            }

    return {'first_number': 15, 'second_number': 7, 'answer': 8}


def generate_custom_multiplication(min_val, max_val):
    """توليد سؤال ضرب مخصص"""
    for _ in range(50):
        num1 = random.randint(min_val, max_val)
        num2 = random.randint(min_val, max_val)

        if num1 == 10 or num2 == 10:
            continue

        first_number = min(num1, num2)
        second_number = max(num1, num2)
        answer = first_number * second_number

        if '10' not in str(answer):
            return {
                'first_number': first_number,
                'second_number': second_number,
                'answer': answer
            }

    return {'first_number': 3, 'second_number': 7, 'answer': 21}


def generate_custom_division(quotients, divisors):
    """توليد سؤال قسمة مخصص"""
    for _ in range(50):
        quotient = random.choice(quotients)
        divisor = random.choice(divisors)

        if quotient == 10 or divisor == 10:
            continue

        dividend = quotient * divisor

        if '10' not in str(dividend) and '10' not in str(quotient):
            return {
                'first_number': divisor,      # المقسوم عليه على اليسار
                'second_number': dividend,    # المقسوم على اليمين
                'answer': quotient
            }

    return {'first_number': 3, 'second_number': 21, 'answer': 7}


def get_operation_symbol(operation):
    """الحصول على رمز العملية"""
    symbols = {
        'addition': '+',
        'subtraction': '-',
        'multiplication': '×',
        'division': '÷'
    }
    return symbols.get(operation, '+')


# ==================== Teacher Views for Student Analytics ====================

@login_required
def student_analytics(request):
    """صفحة إحصائيات التلاميذ للأستاذ"""
    # جلب جميع جلسات التلاميذ
    sessions = StudentSession.objects.all().order_by('-start_time')

    # إحصائيات عامة
    total_sessions = sessions.count()
    completed_sessions = sessions.filter(is_completed=True).count()
    total_students = sessions.values('student_name').distinct().count()

    # إحصائيات حسب المستوى الدراسي
    grade_stats = []
    for grade_value, grade_display in StudentSession.GRADE_CHOICES:
        grade_sessions = sessions.filter(grade=grade_value, is_completed=True)
        if grade_sessions.exists():
            avg_score = grade_sessions.aggregate(avg_score=Avg('total_score'))['avg_score'] or 0
            avg_success_rate = sum([s.success_rate for s in grade_sessions]) / len(grade_sessions)

            grade_stats.append({
                'grade': grade_value,
                'grade_display': grade_display,
                'total_sessions': grade_sessions.count(),
                'avg_score': round(avg_score, 1),
                'avg_success_rate': round(avg_success_rate, 1)
            })

    # إحصائيات حسب مستوى الصعوبة
    difficulty_stats = []
    for difficulty_value, difficulty_display in StudentSession.DIFFICULTY_CHOICES:
        diff_sessions = sessions.filter(difficulty=difficulty_value, is_completed=True)
        if diff_sessions.exists():
            avg_score = diff_sessions.aggregate(avg_score=Avg('total_score'))['avg_score'] or 0
            avg_success_rate = sum([s.success_rate for s in diff_sessions]) / len(diff_sessions)

            difficulty_stats.append({
                'difficulty': difficulty_value,
                'difficulty_display': difficulty_display,
                'total_sessions': diff_sessions.count(),
                'avg_score': round(avg_score, 1),
                'avg_success_rate': round(avg_success_rate, 1)
            })

    # أفضل 10 تلاميذ
    top_students = []
    student_names = sessions.filter(is_completed=True).values_list('student_name', flat=True).distinct()
    for name in student_names:
        student_sessions = sessions.filter(student_name=name, is_completed=True)
        if student_sessions.exists():
            best_session = student_sessions.order_by('-total_score').first()
            avg_success_rate = sum([s.success_rate for s in student_sessions]) / len(student_sessions)

            top_students.append({
                'name': name,
                'best_score': best_session.total_score,
                'avg_success_rate': round(avg_success_rate, 1),
                'total_sessions': student_sessions.count(),
                'grade_display': best_session.get_grade_display(),
                'last_session': best_session.start_time
            })

    top_students = sorted(top_students, key=lambda x: x['best_score'], reverse=True)[:10]

    context = {
        'total_sessions': total_sessions,
        'completed_sessions': completed_sessions,
        'total_students': total_students,
        'grade_stats': grade_stats,
        'difficulty_stats': difficulty_stats,
        'top_students': top_students,
        'recent_sessions': sessions.filter(is_completed=True)[:10]
    }

    return render(request, 'competitions/student_analytics.html', context)


@login_required
def student_session_detail(request, session_id):
    """تفاصيل جلسة تلميذ محددة"""
    session = get_object_or_404(StudentSession, id=session_id)
    responses = StudentResponse.objects.filter(session=session).order_by('created_at')

    # تحليل الأداء
    operation_analysis = {
        'addition': {
            'correct': session.addition_correct,
            'total': session.addition_total,
            'percentage': session.addition_percentage
        },
        'subtraction': {
            'correct': session.subtraction_correct,
            'total': session.subtraction_total,
            'percentage': session.subtraction_percentage
        },
        'multiplication': {
            'correct': session.multiplication_correct,
            'total': session.multiplication_total,
            'percentage': session.multiplication_percentage
        },
        'division': {
            'correct': session.division_correct,
            'total': session.division_total,
            'percentage': session.division_percentage
        }
    }

    context = {
        'session': session,
        'responses': responses,
        'operation_analysis': operation_analysis,
        'strengths': session.strengths,
        'weaknesses': session.weaknesses
    }

    return render(request, 'competitions/student_session_detail.html', context)


@login_required
def export_student_session_excel(request, session_id):
    """تصدير تفاصيل جلسة تلميذ إلى Excel"""
    try:
        session = get_object_or_404(StudentSession, id=session_id)

        # إنشاء workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"جلسة {session.student_name}"

        # تنسيق الخلايا
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        data_font = Font(size=12)

        # العنوان الرئيسي
        ws.merge_cells('A1:D1')
        ws['A1'] = f"تفاصيل جلسة {session.student_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        # معلومات أساسية
        row = 3
        ws[f'A{row}'] = "المعلومات الأساسية"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        basic_info = [
            ("اسم التلميذ", session.student_name),
            ("المستوى الدراسي", session.get_grade_display()),
            ("مستوى الصعوبة", session.get_difficulty_display()),
            ("تاريخ الجلسة", session.created_at.strftime('%Y-%m-%d %H:%M')),
            ("النقاط الإجمالية", session.total_score),
            ("إجمالي الأسئلة", session.total_questions),
            ("الإجابات الصحيحة", session.correct_answers),
            ("معدل النجاح", f"{session.success_rate:.1f}%"),
            ("الوقت المستغرق", f"{session.duration_minutes:.1f} دقيقة"),
        ]

        for label, value in basic_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = data_font
            ws[f'B{row}'].font = data_font
            row += 1

        # النتائج حسب العملية
        row += 2
        ws[f'A{row}'] = "النتائج حسب العملية"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        ws[f'A{row}'] = "العملية"
        ws[f'B{row}'] = "الصحيح"
        ws[f'C{row}'] = "الإجمالي"
        ws[f'D{row}'] = "النسبة"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")

        row += 1
        operations_data = [
            ("الجمع", session.addition_correct, session.addition_total, f"{session.addition_percentage:.1f}%"),
            ("الطرح", session.subtraction_correct, session.subtraction_total, f"{session.subtraction_percentage:.1f}%"),
            ("الضرب", session.multiplication_correct, session.multiplication_total, f"{session.multiplication_percentage:.1f}%"),
            ("القسمة", session.division_correct, session.division_total, f"{session.division_percentage:.1f}%"),
        ]

        for operation, correct, total, percentage in operations_data:
            ws[f'A{row}'] = operation
            ws[f'B{row}'] = correct
            ws[f'C{row}'] = total
            ws[f'D{row}'] = percentage

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = data_font
            row += 1

        # تنسيق العرض
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

        # إعداد الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"جلسة_{session.student_name}_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        messages.error(request, f'خطأ في تصدير Excel: {str(e)}')
        return redirect('competitions:student_session_detail', session_id=session_id)


@login_required
def export_student_session_pdf(request, session_id):
    """تصدير تفاصيل جلسة تلميذ إلى PDF"""
    try:
        session = get_object_or_404(StudentSession, id=session_id)

        # إنشاء PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"جلسة_{session.student_name}_{timezone.now().strftime('%Y-%m-%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # إنشاء المحتوى
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # قائمة العناصر
        elements = []

        # الأنماط
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # وسط
        )

        # العنوان
        title = Paragraph(f"تفاصيل جلسة {session.student_name}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))

        # البيانات الأساسية
        data = [
            ['المعلومات الأساسية', ''],
            ['اسم التلميذ', session.student_name],
            ['المستوى الدراسي', session.get_grade_display()],
            ['مستوى الصعوبة', session.get_difficulty_display()],
            ['تاريخ الجلسة', session.created_at.strftime('%Y-%m-%d %H:%M')],
            ['النقاط الإجمالية', str(session.total_score)],
            ['إجمالي الأسئلة', str(session.total_questions)],
            ['الإجابات الصحيحة', str(session.correct_answers)],
            ['معدل النجاح', f"{session.success_rate:.1f}%"],
            ['الوقت المستغرق', f"{session.duration_minutes:.1f} دقيقة"],
        ]

        table = Table(data, colWidths=[3*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        elements.append(Spacer(1, 30))

        # النتائج حسب العملية
        operations_data = [
            ['العملية', 'الصحيح', 'الإجمالي', 'النسبة'],
            ['الجمع', str(session.addition_correct), str(session.addition_total), f"{session.addition_percentage:.1f}%"],
            ['الطرح', str(session.subtraction_correct), str(session.subtraction_total), f"{session.subtraction_percentage:.1f}%"],
            ['الضرب', str(session.multiplication_correct), str(session.multiplication_total), f"{session.multiplication_percentage:.1f}%"],
            ['القسمة', str(session.division_correct), str(session.division_total), f"{session.division_percentage:.1f}%"],
        ]

        operations_table = Table(operations_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        operations_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(operations_table)

        # بناء PDF
        doc.build(elements)

        # إرجاع المحتوى
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    except Exception as e:
        messages.error(request, f'خطأ في تصدير PDF: {str(e)}')
        return redirect('competitions:student_session_detail', session_id=session_id)